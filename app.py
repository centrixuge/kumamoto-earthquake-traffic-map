#!/usr/bin/env python
# coding: utf-8
"""
熊本地震（2026-07-28）前後の交通量を平常時と比べるダッシュボード。

`fetch_and_prepare.py` が生成した data/*.parquet, data/quake_info.json を
読み込んで表示するだけのビュー層。GDAL依存のgeopandas/shapelyはここでは使わない
（Streamlit Community Cloud等の軽量環境でも動かせるようにするため）。
"""
import io
import itertools
import json
import math
import os
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone

import branca.element as branca_element
import folium
from folium import plugins as folium_plugins
import jinja2
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from streamlit_folium import st_folium

from modules.holidays import WEEKDAY_LABELS
from modules.nexco_text import emergency_lines, emergency_note
from modules import mesh_population, mlit_map_view
from modules.stations import attach_point_code, load_station_master

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
MAX_SELECTED_POINTS = 2
# 地図の大きさと縮尺。st_folium に渡す高さと、folium 側の zoom_start。
# 中心を震源へ寄せる余地がどれだけあるかはこの2つで決まるので、
# 定数として1か所に置く。
# 異常検知（zスコアと地図の色分け）の対象期間。本震からこの長さだけ。
# データの取得は本震+3週間まで伸びるが、その後半はお盆にかかり、
# 地震前の平常時と比べる意味が薄い。
ANOMALY_WINDOW = timedelta(days=7)
MAP_HEIGHT_PX = 496
MAP_ZOOM = 10
# 中心を寄せるときに、観測点マーカーが端で欠けないよう空けておく余白。
# マーカーは最大でも直径28pxなので、その半分に少し足した値。
MAP_EDGE_MARGIN_PX = 22
# 幅は列幅（ブラウザの表示倍率や画面幅で変わる）に追従する。実測414pxだが、
# 狭い画面で観測点が切れないよう、余地の計算では控えめな値を使う。
MAP_MIN_WIDTH_PX = 400
# 時系列図の下余白。日付の目盛りと、横2行までの凡例が入る高さ（実測値）。
CHART_BOTTOM_MARGIN = 85
# 凡例が3行以上になるときに1行あたり足す高さと、1行に並ぶ項目数。
# どちらも実測値（車種別は項目名が長いので1行3項目まで）。
CHART_LEGEND_ROW_PX = 19
CHART_LEGEND_PER_ROW = 3
SELECTION_COLORS = ["red", "green"]
# 車種別ビューで大型車に使う色。選択色（赤・緑）と同系統のまま明るくして、
# 「同じ観測点の別の車種」だと分かるようにする。大型車は小型車の1割程度の
# 水準なので同じ縦軸でも下側に分かれて描かれる。
# 線種で分ける案は、平常時の点線と紛らわしくなるため採らなかった。
SELECTION_ALT_COLORS = ["#ff5fa2", "#7ac70c"]  # ピンク / ライムグリーン
# 観測点マーカーのツールチップに必ず入れる定型句。クリックされた図形が
# 観測点マーカーかどうかを、この文字列で判定する（point_id_from_tooltip）。
# 表示と判定でずれないよう1か所に置く。
POINT_TOOLTIP_HINT = "（クリックで交通量のグラフに表示/解除）"
# 交通量API利用規約（2025-05-12施行）が求める記載。
#   第5条1項: サービス提供時、エンドユーザーが利用の度に確認できる位置に明示
#   第5条2項: 加工して利用する場合の出典
#   第8条2項: 利用者が一切の責任を負う旨の明記
# 文言は規約の記載例に合わせ、3つを1文にまとめて常時表示する。
JARTIC_TERMS_NOTICE = (
    "このサービスは、交通量API 機能を使用していますが、"
    "サービスの内容は国土交通省によって保証されたものではありません。"
    "掲載している交通量は「国土交通省API 機能による交通量(参考値)を加工して作成」"
    "したものです。本サービスの作成・運営について、"
    '<a href="https://github.com/centrixuge" target="_blank">作成者</a>'
    "が一切の責任を負います。"
)

# 時系列図の表示開始時刻（データ保持期間の先頭より後ろにしている）
TIMESERIES_DISPLAY_START = pd.Timestamp("2026-07-27 12:00")
# 表示期間の選択肢。データは復旧期（本震+3週間）まで伸び続けるので、
# 常に最新まで出すと発災直後の変化が横に潰れて読めなくなる。
# 区切りは本震発生時刻を起点にした経過時間で置く（日付で書くと
# 何日目なのかが読み取れないため）。開始はどれも発災前からで
# 自明なので、名前には終端だけを書いている。
# 値は (本震時刻, データの最終時刻) を受け取って (開始, 終了) を返す。
# 本震時刻は quake_info.json から読んだものを渡す（ここで日時を
# 直書きすると二重管理になる）。終了がデータより後ろでも Plotly 側で
# 右側が空くだけなので、データがそこまで伸びるまでの間も同じ窓を出せる。
def _day_ticks(x_range):
    """x軸の目盛りを日の変わり目に置き、8/4(火) の形で曜日まで出す。

    plotly の tickformat は d3-time-format なので、日本語の曜日を
    出せない（%a は Tue になる）。目盛りの位置と文字列を自前で作る。
    平常時が日区分（曜日）ごとに違うため、どの曜日と比べているのかを
    軸から読めるようにしておきたい。
    """
    start, end = pd.Timestamp(x_range[0]), pd.Timestamp(x_range[1])
    # 表示幅が広いときは1日おきにして、目盛りが詰まりすぎないようにする
    step = 1 if (end - start) <= pd.Timedelta(days=9) else 2
    first = start.normalize()
    if first < start:
        first += pd.Timedelta(days=1)
    days = pd.date_range(first, end, freq=f"{step}D")
    text = [f"{d.month}/{d.day}({WEEKDAY_LABELS[d.weekday()]})" for d in days]
    return list(days), text


def _since_quake(days: int):
    """本震から days 日後を含む日の24時までを返す。

    本震は16:27なので厳密に days 日後で切ると日の途中で終わり、
    右端が半端な位置になる。その時点を含む日の終わり（翌日0時）まで
    伸ばして、日の区切りで終わるようにしている。
    """
    return lambda quake, last: (
        TIMESERIES_DISPLAY_START,
        (quake + pd.Timedelta(days=days)).normalize() + pd.Timedelta(days=1),
    )


TIMESERIES_RANGES = {
    "発災後3日間": _since_quake(3),
    "発災後1週間": _since_quake(7),
    "発災後2週間": _since_quake(14),
    "発災後3週間": _since_quake(21),
    "最新3日間": lambda quake, last: (last - pd.Timedelta(days=3), last),
}
# 既定は発災後1週間。発災直後の落ち込みと戻り始めが1枚で読める。
TIMESERIES_DEFAULT_RANGE = "発災後1週間"

# 時系列ビューの切り替え。実績の既定は5分間値、平常時はいずれも1時間値
# （同じ日区分の各8日分）から求めている。
TIMESERIES_VIEWS = {
    "5分間値": {
        "file": "observations.parquet",
        "unit_label": "5分間値",
        "note": (
            "平常時は1時間値から求め、単位を合わせるため1/12しています。"
            "そのため帯は「平常時の1時間あたり交通量の日々のばらつき÷12」であり、"
            "5分間値そのもののばらつきではありません（帯は狭めに出ます）。"
            "異常検知の判定は1時間値ベースで行っています。"
        ),
    },
    "1時間値": {
        "file": "observations_hourly.parquet",
        "unit_label": "1時間値",
        "note": "実績・平常時とも1時間値どうしの比較で、異常検知の判定もこの粒度で行っています。",
    },
    "車種別・5分間値": {
        "file": "observations.parquet",
        "unit_label": "5分間値",
        "series": "vehicle",
        "note": (
            "同じ図の中に小型車（濃い色）と大型車（明るい色）を5分間値で並べています。"
            "**大型車は5分あたり中央値4台（0台が14%）と粒度が粗く**、"
            "線がぎざぎざになります。形の変化を細かく追いたいとき向けで、"
            "水準の比較は「車種別・1時間値」のほうが読みやすいです。"
            "平常時は1時間値から求めた値を1/12しているため、帯は実際のばらつきより狭く出ます"
            "（台数の小さい大型車ではこの影響がより大きくなります）。"
            "異常検知の判定は車種を合計した1時間値で行っており、この図の判定は変わりません。"
        ),
    },
    "車種別・1時間値": {
        "file": "observations_hourly.parquet",
        "unit_label": "1時間値",
        "series": "vehicle",
        "note": (
            "同じ図の中に小型車（濃い色）と大型車（明るい色）を並べています。"
            "大型車は小型車の1割程度の水準なので、同じ縦軸でも下側に分かれて描かれます。"
            "線種で分ける案は平常時の点線と紛らわしくなるため色で分けています。"
            "異常検知の判定は車種を合計した1時間値で行っており、この図の判定は変わりません。"
            "APIは車種判別不能の台数も返しますが、アーカイブ全期間で1台しかないため扱っていません。"
        ),
    },
}

# 時系列図に描く系列。既定は上り・下りの合計、「車種別」は小型・大型に分ける。
# 大型車は小型車の1割程度なので、同じ軸に載せず図を分ける。
# 1図＝1方向。各図に描く系列を (列のサフィックス, 凡例に足す語, 色の種別) で持つ。
# 色の種別 "main" は選択色そのもの、"alt" は同系統の明るい色（大型車用）。
SERIES_TOTAL = [
    ("上り", [("up", "", "main")]),
    ("下り", [("down", "", "main")]),
]
SERIES_VEHICLE = [
    ("上り", [("up_small", " 小型", "main"), ("up_large", " 大型", "alt")]),
    ("下り", [("down_small", " 小型", "main"), ("down_large", " 大型", "alt")]),
]

# 通行規制の日時はJSTの壁時計時刻（naive）で保存されている。Streamlit Cloud等の
# UTCサーバーでdatetime.now()をそのまま使うと「終了済みか」の判定がずれるため、
# fetch_and_prepare.py と同様にJST固定の「今」を使う。
JST = timezone(timedelta(hours=9))


def _now_jst() -> datetime:
    return datetime.now(JST).replace(tzinfo=None)

st.set_page_config(
    page_title="熊本地震・交通量変化ダッシュボード",
    layout="wide",
)


@st.cache_data(ttl=300)
def load_observations(filename: str = "observations.parquet") -> pd.DataFrame:
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_parquet(path)
    df["datetime"] = pd.to_datetime(df["datetime"])
    return df


@st.cache_data(ttl=300)
def load_quake_info() -> dict:
    with open(os.path.join(DATA_DIR, "quake_info.json"), encoding="utf-8") as f:
        return json.load(f)


@st.cache_data(ttl=300)
def load_regulations() -> list:
    path = os.path.join(DATA_DIR, "regulations.json")
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f).get("items", [])


@st.cache_data(ttl=300)
def load_mlit_regulations() -> dict:
    """
    直轄国道（国が管理する国道）の規制。熊本県「防災情報くまもと」の
    通行規制情報は県・市町村が管理する道路が対象で（県が管理する補助国道も
    含む。国道219号など）、国道57号のような
    直轄国道は載らない。そのため熊本河川国道事務所の公表PDFから
    手作業で転記したものを別ファイルで持つ。
    """
    path = os.path.join(DATA_DIR, "mlit_regulations.json")
    if not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8") as f:
        return json.load(f)


@st.cache_data(ttl=300)
def load_nexco_regulations() -> dict:
    """
    高速道路（NEXCO西日本が管理する有料道路）の規制。県のポータルにも
    熊本河川国道事務所のPDFにも載らないため、NEXCO西日本が公表する
    「お知らせ」PDFから手作業で転記したものを別ファイルで持つ。
    """
    path = os.path.join(DATA_DIR, "nexco_regulations.json")
    if not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8") as f:
        return json.load(f)


@st.cache_data(ttl=300)
def regulation_archive_start() -> str:
    """
    通行規制のアーカイブに記録が残っている最も古い時点（first_seen の最小値）。

    これより前に解除された規制は、県管理道路であってもアーカイブに存在しない。
    本震（7/28 16:27）より後に収集を始めたため、発災直後だけ規制されて
    すぐ解除された区間は取りこぼしている。その境目を画面に出すために使う。
    """
    path = os.path.join(DATA_DIR, "archive", "regulations_archive.json")
    if not os.path.exists(path):
        return ""
    with open(path, encoding="utf-8") as f:
        items = json.load(f).get("items", {})
    seen = [v.get("first_seen") for v in items.values() if v.get("first_seen")]
    if not seen:
        return ""
    return pd.Timestamp(min(seen)).strftime("%Y-%m-%d %H:%M")


def mlit_regulations_for_point(mlit: dict, point_code) -> list:
    """指定した観測点コードに掛かっている直轄国道の規制を返す。"""
    if not mlit or point_code is None or pd.isna(point_code):
        return []
    code = str(point_code)
    return [
        item for item in mlit.get("items", [])
        if code in (item.get("affected_point_codes") or [])
    ]


@st.cache_data(ttl=300)
def load_station_master_cached() -> dict:
    """常時観測点コードと緯度経度の対応（fetch_and_prepare.py が生成）。"""
    return load_station_master(os.path.join(DATA_DIR, "stations.json"))


@st.cache_data(ttl=300)
def load_traffic_archive(filename: str) -> pd.DataFrame:
    """
    交通量の恒久アーカイブ（5分値／1時間値）を読む。
    アーカイブにはコード列を持たない時期のデータも含まれるので、
    観測点マスタからJARTICの常時観測点コードを付け直して先頭列に置く。
    """
    path = os.path.join(DATA_DIR, "archive", filename)
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_parquet(path)
    df["datetime"] = pd.to_datetime(df["datetime"])
    df = attach_point_code(df, load_station_master_cached())
    cols = ["point_code"] + [c for c in df.columns if c != "point_code"]
    return df[cols].sort_values(["datetime", "point_code"]).reset_index(drop=True)


@st.cache_data(ttl=300)
def load_regulations_archive() -> tuple:
    """
    通行規制の恒久アーカイブを、CSVにしやすい2つの表に開く。
      - 規制一覧: 1件1行（経路の座標列は列数が膨大になるので含めない）
      - 状態変化履歴: 規制内容・終了日時などが変わった時点ごとに1行
    """
    path = os.path.join(DATA_DIR, "archive", "regulations_archive.json")
    if not os.path.exists(path):
        return pd.DataFrame(), pd.DataFrame()
    with open(path, encoding="utf-8") as f:
        items = (json.load(f).get("items") or {})

    rows, hist_rows = [], []
    for key, rec in items.items():
        rows.append({
            "regulation_key": key,
            "route_name": rec.get("route_name"),
            "region": rec.get("region"),
            "content": rec.get("content"),
            "reason_type": rec.get("reason_type"),
            "reason_detail": rec.get("reason_detail"),
            "start_timestamp": rec.get("start_timestamp"),
            "end_timestamp": rec.get("end_timestamp"),
            "length_km": rec.get("length_km"),
            "start_lat": rec.get("start_lat"), "start_lon": rec.get("start_lon"),
            "end_lat": rec.get("end_lat"), "end_lon": rec.get("end_lon"),
            "first_seen": rec.get("first_seen"),
            "last_seen": rec.get("last_seen"),
            "still_listed": rec.get("still_listed"),
            "path_points": len(rec.get("path") or []),
        })
        for h in rec.get("history") or []:
            hist_rows.append({
                "regulation_key": key,
                "route_name": rec.get("route_name"),
                "observed_at": h.get("observed_at"),
                "content": h.get("content"),
                "end_timestamp": h.get("end_timestamp"),
                "reason_type": h.get("reason_type"),
                "reason_detail": h.get("reason_detail"),
                "length_km": h.get("length_km"),
            })

    regs = pd.DataFrame(rows).sort_values("start_timestamp", ascending=False)
    hist = pd.DataFrame(hist_rows)
    if not hist.empty:
        hist = hist.sort_values(["observed_at", "route_name"])
    return regs.reset_index(drop=True), hist.reset_index(drop=True)


@st.cache_data(ttl=300)
def to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Excelでそのまま開けるようBOM付きUTF-8にする。"""
    return df.to_csv(index=False).encode("utf-8-sig")


# ダウンロードするCSVの列定義。「データセット名」はダウンロードするCSVのファイル名と
# 対応させている。実データの列と食い違っていないかは build_data_dictionary で検査する。
_TRAFFIC_COLUMNS = [
    ("point_code", "文字列", "JARTICの常時観測点コード。観測点を一意に識別する公式のID"),
    ("road_type", "文字列", "「1」＝高速自動車国道、「3」＝一般国道（JARTICの道路種別。この2値のみ）。常時観測点コードからは判別できないためAPIの値を保持している。地図では高速自動車国道を■、一般国道を●で描き分けている"),
    ("lon", "度（EPSG:4326）", "観測点の経度"),
    ("lat", "度（EPSG:4326）", "観測点の緯度"),
    ("datetime", "日時（JST）", "観測時刻。5分間値は5分刻み、1時間値は毎時0分（その時刻から始まる区間の集計値）"),
    ("traffic_up", "台", "上り方向の合計交通量（小型＋大型＋車種判別不能）。いずれかが欠測なら空欄"),
    ("traffic_down", "台", "下り方向の合計交通量（同上）"),
    ("traffic_up_small", "台", "上り・小型車の交通量"),
    ("traffic_up_large", "台", "上り・大型車の交通量"),
    ("traffic_up_unidentified", "台", "上り・車種を判別できなかった交通量"),
    ("traffic_down_small", "台", "下り・小型車の交通量"),
    ("traffic_down_large", "台", "下り・大型車の交通量"),
    ("traffic_down_unidentified", "台", "下り・車種を判別できなかった交通量"),
]

_OBSERVATION_COLUMNS = [
    ("point_code", "文字列", "JARTICの常時観測点コード"),
    ("road_type", "文字列", "「1」＝高速自動車国道、「3」＝一般国道（JARTICの道路種別。この2値のみ）。常時観測点コードからは判別できないためAPIの値を保持している。地図では高速自動車国道を■、一般国道を●で描き分けている"),
    ("point_id", "文字列", "内部の結合キー（経度_緯度を6桁に丸めた文字列）。観測点の識別にはpoint_codeを使ってください"),
    ("point_lon", "度（EPSG:4326）", "観測点の経度（6桁に丸め）"),
    ("point_lat", "度（EPSG:4326）", "観測点の緯度（6桁に丸め）"),
    ("datetime", "日時（JST）", "観測時刻（1時間値なので毎時0分）"),
    ("daytype", "文字列", "日区分（月/火/水/木/金/土/日祝）。03:00起点の日付で判定し、祝日は曜日によらず日祝に入れる。平常時は同じ日区分どうしで比べる"),
    ("hour", "0〜23", "時刻の「時」。平常時の平均・標準偏差は日区分×この単位で求めている"),
    ("traffic_up", "台/時", "上り方向の実績交通量（小型＋大型）"),
    ("traffic_down", "台/時", "下り方向の実績交通量（小型＋大型）"),
    ("traffic_up_small", "台/時", "上り・小型車の実績交通量"),
    ("traffic_up_large", "台/時", "上り・大型車の実績交通量"),
    ("traffic_down_small", "台/時", "下り・小型車の実績交通量"),
    ("traffic_down_large", "台/時", "下り・大型車の実績交通量"),
    ("baseline_mean_up", "台/時", "平常時（同じ日区分の8日分）の上り交通量の平均"),
    ("baseline_std_up", "台/時", "同じ母集団での上り交通量の標準偏差"),
    ("baseline_mean_down", "台/時", "平常時の下り交通量の平均"),
    ("baseline_std_down", "台/時", "平常時の下り交通量の標準偏差"),
    ("baseline_mean_up_small", "台/時", "平常時の上り・小型車の平均"),
    ("baseline_std_up_small", "台/時", "同じ母集団での上り・小型車の標準偏差"),
    ("baseline_mean_up_large", "台/時", "平常時の上り・大型車の平均"),
    ("baseline_std_up_large", "台/時", "同じ母集団での上り・大型車の標準偏差"),
    ("baseline_mean_down_small", "台/時", "平常時の下り・小型車の平均"),
    ("baseline_std_down_small", "台/時", "同じ母集団での下り・小型車の標準偏差"),
    ("baseline_mean_down_large", "台/時", "平常時の下り・大型車の平均"),
    ("baseline_std_down_large", "台/時", "同じ母集団での下り・大型車の標準偏差"),
    ("z_up", "無次元", "上りのzスコア =（実績 − 平常時平均）÷ 標準偏差。車種を合計した値で計算する。標準偏差が0や極小のときは下限でクリップ"),
    ("z_down", "無次元", "下りのzスコア（同上）"),
    ("is_post_quake", "真偽値", "本震（2026-07-28 16:27）以降の時刻かどうか"),
    ("is_anomaly", "真偽値", "異常と判定したか。is_post_quakeがTrueかつ|z_up|または|z_down|が2以上"),
    ("distance_km_from_epicenter", "km", "観測点と本震の震源との大圏距離（参考値。異常検知の判定には使っていない）"),
]

_REGULATION_COLUMNS = [
    ("regulation_key", "文字列", "規制の識別キー。路線名・地域・区間の始終点座標・規制開始日時を連結したもの（規制内容や終了日時は途中で変わるため含めない）"),
    ("route_name", "文字列", "路線名"),
    ("region", "文字列", "振興局などの地域区分"),
    ("content", "文字列", "最新の規制内容（全面通行止め／片側交互通行止め／車両通行止め／解除 など）"),
    ("reason_type", "文字列", "規制の原因区分（災害／工事／事故／その他）"),
    ("reason_detail", "文字列", "原因の詳細（道路損壊など）。空欄のことも多い"),
    ("start_timestamp", "日時（JST）", "規制の開始日時。本震（2026-07-28 16:27）以降かどうかで地震起因かを切り分けている"),
    ("end_timestamp", "日時（JST）", "規制の終了日時（実績または予定）。未定なら空欄"),
    ("length_km", "km", "規制区間の延長（ポータルの申告値）"),
    ("start_lat", "度（EPSG:4326）", "規制区間の始点の緯度"),
    ("start_lon", "度（EPSG:4326）", "規制区間の始点の経度"),
    ("end_lat", "度（EPSG:4326）", "規制区間の終点の緯度"),
    ("end_lon", "度（EPSG:4326）", "規制区間の終点の経度"),
    ("first_seen", "日時（JST）", "この規制をアーカイブで最初に確認した日時"),
    ("last_seen", "日時（JST）", "最後に確認した日時。still_listedがFalseならこの時点以降に一覧から消えた"),
    ("still_listed", "真偽値", "最新の取得時点でポータルの一覧に載っていたか。Falseは解除等で削除されたことを示す"),
    ("path_points", "個数", "地図描画用にOSRMで道路網へスナップした経路の座標点数（経路そのものはCSVには含めない）"),
]

_REGULATION_HISTORY_COLUMNS = [
    ("regulation_key", "文字列", "規制の識別キー（一覧CSVのregulation_keyと対応）"),
    ("route_name", "文字列", "路線名"),
    ("observed_at", "日時（JST）", "この状態を観測した日時"),
    ("content", "文字列", "その時点の規制内容"),
    ("end_timestamp", "日時（JST）", "その時点で示されていた終了日時"),
    ("reason_type", "文字列", "その時点の原因区分"),
    ("reason_detail", "文字列", "その時点の原因詳細"),
    ("length_km", "km", "その時点の規制区間延長"),
]

# (CSVファイル名, Excelのシート名, データセット名, 列定義)
# シート名はExcelの制約（31文字以内・ : \ / ? * [ ] を含まない）に収まるよう短くしている。
DATA_DICTIONARY = [
    ("kumamoto_traffic_5min_archive.csv", "5分間交通量", "5分間交通量（アーカイブ全期間）", _TRAFFIC_COLUMNS),
    ("kumamoto_traffic_hourly_archive.csv", "1時間交通量", "1時間交通量（アーカイブ全期間）", _TRAFFIC_COLUMNS),
    ("kumamoto_road_regulations_archive.csv", "通行規制_一覧", "通行規制 一覧", _REGULATION_COLUMNS),
    ("kumamoto_road_regulations_history.csv", "通行規制_履歴", "通行規制 状態変化の履歴", _REGULATION_HISTORY_COLUMNS),
    ("kumamoto_observations_hourly.csv", "異常検知_入力データ", "異常検知の入力データ（1時間値）", _OBSERVATION_COLUMNS),
]


def build_data_dictionary(actual_columns: dict) -> tuple:
    """
    列定義書を1枚の表にする。あわせて、実際のCSVの列と定義が食い違っていないかを
    検査し、その内容を「読んで意味が分かる備考文」として返す。

    食い違いは異常ではなく、次の仕組みで一時的に起こりうる:
      - 列定義書はアプリのコード内にあり、デプロイした時点で最新になる
      - CSVの中身は「最後にデータを生成した時点」のもの（6時間ごとの自動実行 or 手動実行）
    そのため列を追加した直後は、次のデータ生成までの間だけ両者がずれる。
    """
    rows = []
    only_in_dict, only_in_csv = {}, {}
    for fname, _sheet, dataset, columns in DATA_DICTIONARY:
        documented = [c for c, _, _ in columns]
        actual = actual_columns.get(fname)
        if actual is not None:
            stale = [c for c in documented if c not in actual]
            extra = [c for c in actual if c not in documented]
            if stale:
                only_in_dict[fname] = stale
            if extra:
                only_in_csv[fname] = extra
        for order, (col, unit, desc) in enumerate(columns, start=1):
            rows.append({
                "csv_file": fname,
                "dataset": dataset,
                "column_order": order,
                "column": col,
                "unit_or_type": unit,
                "description": desc,
                "in_actual_csv": (col in actual) if actual is not None else None,
            })

    notes = []
    if only_in_dict:
        detail = "、".join(
            f"`{f}` の {', '.join(f'`{c}`' for c in cols)}"
            for f, cols in only_in_dict.items()
        )
        notes.append(
            f"**備考: 定義書に載っているのに、いま配布中のCSVにまだ入っていない列があります**（{detail}）。"
            "エラーではありません。列定義書はアプリのコードに書かれていてデプロイと同時に新しくなる一方、"
            "CSVの中身は「最後にデータを生成した時点」のものなので、列を増やした直後は"
            "次のデータ生成（6時間ごとの自動実行、または `python fetch_and_prepare.py`）が走るまでの間だけ"
            "両者がずれます。このずれがある間だけ、列定義書に「配布中CSVに存在」列が追加され、"
            "該当する列が `×` になります（全て揃っているときはこの列自体を省いています）。"
        )
        if any("point_code" in cols for cols in only_in_dict.values()):
            notes.append(
                "**`point_code` について**: この列は観測点マスタ（`data/stations.json`：常時観測点コードと"
                "緯度経度の対応表）を使って付けています。マスタが用意される前に生成されたデータには"
                "この列自体が存在しないため、いったん定義書だけに載る状態になります。"
                "次のデータ生成でマスタから付与され、CSVにも入ります。"
                "それまでは `lon` / `lat` の組が観測点の識別子として使えます。"
            )
    if only_in_csv:
        detail = "、".join(
            f"`{f}` の {', '.join(f'`{c}`' for c in cols)}"
            for f, cols in only_in_csv.items()
        )
        notes.append(
            f"**備考: CSVに入っているのに、定義書でまだ説明していない列があります**（{detail}）。"
            "データ側に新しい列が増えて、定義書の更新が追いついていない状態です。"
        )
    return pd.DataFrame(rows), notes


def _plain_text(text: str) -> str:
    """備考文のMarkdown記法（**強調**・`コード`）を落としてExcelに書ける素の文にする。"""
    return text.replace("**", "").replace("`", "")


# 列定義シートのヘッダ。Excelで人が読む資料なので日本語にする。
_DICT_SHEET_HEADERS = [
    ("列番号", 8),
    ("列名", 30),
    ("単位・型", 18),
    ("説明", 90),
]
_DICT_PRESENCE_HEADER = ("配布中CSVに存在", 20)


def dictionary_has_missing_column(dict_df: pd.DataFrame) -> bool:
    """
    定義書にあるのに配布中CSVに無い列があるか。全て揃っているときは
    「配布中CSVに存在」列が全て○になって情報量がないので、この判定で列自体を省く。
    """
    if dict_df.empty or "in_actual_csv" not in dict_df.columns:
        return False
    return bool(dict_df["in_actual_csv"].eq(False).any())


@st.cache_data(ttl=300)
def to_dictionary_xlsx_bytes(dict_df: pd.DataFrame, notes: tuple) -> bytes:
    """
    列定義書をExcelブックにする。1シート＝1データセット（＝1CSVファイル）とし、
    先頭に全体の目次と備考をまとめた「はじめに」シートを置く。
    「配布中CSVに存在」列は食い違いがあるときだけ追加する。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    wrap_top = Alignment(vertical="top", wrap_text=True)
    top = Alignment(vertical="top")

    show_presence = dictionary_has_missing_column(dict_df)
    headers = _DICT_SHEET_HEADERS + ([_DICT_PRESENCE_HEADER] if show_presence else [])

    wb = Workbook()
    intro = wb.active
    intro.title = "はじめに"

    intro["A1"] = "熊本地震・交通量変化ダッシュボード CSV列定義書"
    intro["A1"].font = Font(bold=True, size=14)
    intro["A2"] = f"作成日時: {_now_jst():%Y-%m-%d %H:%M} (JST)"
    intro["A3"] = "公開URL: https://kumamoto-earthquake-traffic-map.streamlit.app/"
    intro["A4"] = "データセットごとにシートを分けています。下の表のシート名から移動してください。"

    row = 6
    for i, label in enumerate(["シート名", "CSVファイル名", "データセット", "列数"]):
        cell = intro.cell(row=row, column=i + 1, value=label)
        cell.font = header_font
        cell.fill = header_fill
    for fname, sheet, dataset, columns in DATA_DICTIONARY:
        row += 1
        intro.cell(row=row, column=1, value=sheet)
        intro.cell(row=row, column=2, value=fname)
        intro.cell(row=row, column=3, value=dataset)
        intro.cell(row=row, column=4, value=len(columns))

    if notes:
        row += 2
        cell = intro.cell(row=row, column=1, value="備考")
        cell.font = header_font
        for note in notes:
            row += 1
            c = intro.cell(row=row, column=1, value=_plain_text(note))
            c.alignment = wrap_top
    for col, width in zip("ABCD", (22, 42, 34, 8)):
        intro.column_dimensions[col].width = width

    for fname, sheet, dataset, _columns in DATA_DICTIONARY:
        ws = wb.create_sheet(sheet)
        ws["A1"] = dataset
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = f"CSVファイル名: {fname}"
        for i, (label, width) in enumerate(headers):
            cell = ws.cell(row=4, column=i + 1, value=label)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(i + 1)].width = width

        sub = dict_df[dict_df["csv_file"] == fname]
        for offset, (_, r) in enumerate(sub.iterrows()):
            out = 5 + offset
            values = [r["column_order"], r["column"], r["unit_or_type"], r["description"]]
            if show_presence:
                in_csv = r["in_actual_csv"]
                values.append(
                    "-" if in_csv is None else ("○" if in_csv else "×（次回生成時に追加）")
                )
            for i, v in enumerate(values):
                c = ws.cell(row=out, column=i + 1, value=v)
                c.alignment = wrap_top if i == 3 else top
        # ヘッダ行を固定して、列数の多いデータセットでもスクロールしやすくする
        ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@st.cache_data(ttl=300)
def build_point_summary(post: pd.DataFrame, observations: pd.DataFrame = None) -> pd.DataFrame:
    """
    観測点ごとの異常度をまとめる。

    地震後の交通量が全て欠測の観測点は max_abs_z が NaN になる（zスコアが
    定義できない）。これは「異常なし」ではなく「地震後のデータがない」という
    別の状態なので、地図側で区別して描けるよう last_observed_at（最後に値が
    あった時刻）も持たせる。実際に地震発生と同時に配信が止まった観測点がある。
    """
    if post.empty:
        return pd.DataFrame(columns=[
            "point_id", "point_code", "road_type", "point_lon", "point_lat",
            "max_abs_z", "n_anomaly", "distance_km", "last_observed_at",
        ])
    summary = (
        post.groupby(["point_id", "point_lon", "point_lat"])
        .apply(
            lambda g: pd.Series({
                # 常時観測点コードは観測点ごとに一意なので、先頭の値をそのまま採る
                "point_code": (
                    g["point_code"].dropna().iloc[0]
                    if "point_code" in g.columns and g["point_code"].notna().any()
                    else None
                ),
                # 道路種別も観測点ごとに固定（1:高速自動車国道 / 3:一般国道）。
                # 地図のマーカー形状を分けるのに使う。
                "road_type": (
                    g["road_type"].dropna().iloc[0]
                    if "road_type" in g.columns and g["road_type"].notna().any()
                    else None
                ),
                "max_abs_z": max(g["z_up"].abs().max(), g["z_down"].abs().max()),
                "n_anomaly": int(g["is_anomaly"].sum()),
                "distance_km": g["distance_km_from_epicenter"].iloc[0],
            }),
            include_groups=False,
        )
        .reset_index()
    )

    # 最後に値があった時刻は地震前も含めて見る必要があるので observations 全体から求める
    src = observations if observations is not None and not observations.empty else post
    has_value = src["traffic_up"].notna() | src["traffic_down"].notna()
    last_seen = (
        src.loc[has_value].groupby("point_id")["datetime"].max().rename("last_observed_at")
    )
    summary = summary.merge(last_seen, on="point_id", how="left")

    return summary.sort_values("max_abs_z", ascending=False).reset_index(drop=True)


# 初期表示で選んでおく観測点（JARTICの常時観測点コード）。
# 異常度の最上位を出すと日によって入れ替わってしまうので、
# 見せたい地点を固定しておく。データに無ければ最上位にフォールバックする。
DEFAULT_POINT_CODE = "9110040"


def _default_selection(point_summary: pd.DataFrame) -> list:
    """初期選択の観測点IDを返す。"""
    if point_summary.empty:
        return []
    if "point_code" in point_summary.columns:
        hit = point_summary[point_summary["point_code"].astype(str) == DEFAULT_POINT_CODE]
        if not hit.empty:
            return [hit["point_id"].iloc[0]]
    return [point_summary["point_id"].iloc[0]]


FULL_CLOSURE_CONTENTS = {"全面通行止め", "車両通行止め"}


def _point_radius(max_abs_z, max_z: float, selected: bool = False) -> float:
    """
    観測点マーカーの半径。異常度に比例させる。欠測（NaN）は最小サイズ。
    ▲の表示位置を丸の外側に置くためにも使うので、計算を1か所にまとめている。
    """
    frac = 0.0 if pd.isna(max_abs_z) else max_abs_z / max_z
    return (14 if selected else 11) + 12 * frac


# 高速自動車国道は■、一般国道は●で描く。色は異常度に使っているので、
# 道路の種別は形で分ける（色を増やすと異常度のグラデーションが読めなくなる）。
ROAD_TYPE_SQUARE = "1"    # 高速自動車国道
ROAD_TYPE_GENERAL = "3"   # 一般国道
ROAD_TYPE_LABELS = {"1": "高速自動車国道", "3": "一般国道"}


def _is_square_point(road_type) -> bool:
    return str(road_type) == ROAD_TYPE_SQUARE


def _road_type_counts(point_summary: pd.DataFrame) -> tuple:
    """(一般国道, 高速自動車国道, 種別不明) の点数を返す。

    種別が取れていない観測点を一般国道に混ぜないよう別に数える。
    アーカイブには道路種別の列を持たない時期のデータがあり、
    配信が止まって観測点マスタにも残っていない地点（9310183 など）は
    種別が引けない。そういう点まで「一般国道」と表示すると、
    根拠のない分類を出すことになる。
    """
    if "road_type" not in point_summary.columns:
        return 0, 0, len(point_summary)
    rt = point_summary["road_type"].astype("string")
    n_express = int((rt == ROAD_TYPE_SQUARE).sum())
    n_general = int((rt == ROAD_TYPE_GENERAL).sum())
    return n_general, n_express, len(point_summary) - n_general - n_express


def _rgba(hex_color: str, alpha: float) -> str:
    """#rrggbb を rgba() にする。■はdivで描くため、塗りだけ透過させたい。"""
    h = hex_color.lstrip("#")
    r, g, b = (int(h[i:i + 2], 16) for i in (0, 2, 4))
    return f"rgba({r},{g},{b},{alpha})"


def _severity_color(frac: float) -> str:
    """
    0(平常)〜1(最大異常)のfracを寒色系（薄い水色〜濃い紺）の16進カラーに変換する。
    通行規制の色（赤系）と見分けやすいよう、観測点側はあえて寒色にしている。
    """
    frac = max(0.0, min(1.0, frac))
    low = (222, 235, 247)
    high = (8, 48, 107)
    rgb = [int(low[i] + (high[i] - low[i]) * frac) for i in range(3)]
    return "#{:02x}{:02x}{:02x}".format(*rgb)


def _parse_reg_time(value) -> datetime:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def _regulation_is_ended(reg: dict, now: datetime) -> bool:
    """規制が既に解除済みかどうかを判定する。"""
    if reg.get("content") == "解除":
        return True
    end_dt = _parse_reg_time(reg.get("end_timestamp"))
    if end_dt is None:
        return False
    return end_dt < now


def _regulation_is_post_quake(reg: dict, quake_at: datetime) -> bool:
    """
    規制の開始日時が本震発生以降かどうか。データには2020年7月豪雨のように
    今回の地震と無関係な長期規制も多数含まれるため、これで切り分ける。
    開始日時が読めないものは、地震起因と誤認させない側（=発災前）に寄せる。
    """
    start_dt = _parse_reg_time(reg.get("start_timestamp"))
    if start_dt is None:
        return False
    return start_dt >= quake_at


def _x_circle_icon(color: str, ended: bool, size: int = 18) -> folium.DivIcon:
    """
    PDFの別添図と同じ ⊗（丸の中に×）の印を作る。

    観測点マーカーが塗りつぶしの丸なので、規制の地点まで丸にすると見分けが
    つかない。PDFの凡例で使われている記号に合わせて、丸の中に×を描く。
    解除済みは外周を点線にする（線の実線／破線と同じ意味）。

    クリックはこの印を透過させて、下にある観測点マーカーに通す
    （markerPane は overlayPane より上にあるため、透過させないと
     観測点の選択を横取りしてしまう）。
    """
    # 線の太さと余白はサイズに比例させる（小さくしても潰れないように）
    sw = round(size * 0.11, 2)
    r = size / 2 - sw
    c = size / 2
    d = r * 0.62
    dash = f' stroke-dasharray="{round(sw * 1.25, 2)},{round(sw * 1.1, 2)}"' if ended else ""
    html = (
        f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}">'
        f'<circle cx="{c}" cy="{c}" r="{r}" fill="#ffffff" fill-opacity="0.9" '
        f'stroke="{color}" stroke-width="{sw}"{dash}/>'
        f'<path d="M{c - d} {c - d} L{c + d} {c + d} M{c + d} {c - d} L{c - d} {c + d}" '
        f'stroke="{color}" stroke-width="{sw}" stroke-linecap="round"/>'
        "</svg>"
    )
    return folium.DivIcon(
        html=html, icon_size=(size, size), icon_anchor=(size // 2, size // 2),
        class_name="reg-x-mark",
    )


def _epicenter_icon(size: int = 30) -> folium.DivIcon:
    """
    震源の印。気象庁の推計震度分布図と同じ、青い × で描く。

    以前は青いピン（星）だったが、ピンは下端が指す位置になるうえ地図の
    上に大きく張り出すため、震源の位置そのものが読みにくかった。
    気象庁の図と同じ記号にすれば、あの図を見慣れている人にはそれだけで
    震源だと分かる。

    規制の ⊗ とは、丸が無いこと・色が青であること・線が太いことで
    見分けられる。
    """
    sw = round(size * 0.13, 2)
    d = size / 2 - sw
    c = size / 2
    html = (
        f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}">'
        f'<path d="M{c - d} {c - d} L{c + d} {c + d} M{c + d} {c - d} '
        f'L{c - d} {c + d}" stroke="#0000ff" stroke-width="{sw}" '
        'stroke-linecap="butt"/>'
        "</svg>"
    )
    return folium.DivIcon(
        html=html, icon_size=(size, size), icon_anchor=(size // 2, size // 2),
    )


def _emergency_access_html(item: dict) -> str:
    """
    「一般車両は通行止めだが、緊急車両は通れる」区間をツールチップに出す。

    通行止めの区間まるごとではなく、その一部だけのことがある。線を分けて
    描くと通行止めの線と重なって読みにくいうえ、PAなど端点の座標を
    別に用意する必要もあるので、属性として文字で出す。
    文言は modules/nexco_text.py に置き、新しい報を取り込むときのPRの
    本文と同じものを使う（別々に書くと見た内容と地図がずれる）。
    """
    lines = emergency_lines(item)
    if not lines:
        return ""
    colors = ["#c05621", "#8a5a2b", "#8a5a2b", "#777"]
    out = ""
    for i, line in enumerate(lines):
        color = colors[min(i, len(colors) - 1)]
        out += f'<span style="color:{color};">{line}</span><br>'
    return out


def emergency_access_note(nexco: dict) -> str:
    """地図の下に出す1行。線を見ただけでは分からないので、ここでも触れる。"""
    return emergency_note(nexco)


def _regulation_style(reg: dict, now: datetime, quake_at: datetime) -> dict:
    """
    線のスタイルを決める。色＝規制の区分、破線＝解除済み、で意味を分けている。
      赤     : 今回の地震以降に始まった全面/車両通行止め（最重要。×印つき）
      橙     : 今回の地震以降に始まったその他の規制（片側交互通行止めなど）
      青灰色 : 地震前から続く規制（工事・過去の災害など。今回の地震とは無関係）
    """
    ended = _regulation_is_ended(reg, now)
    dash = "6,8" if ended else None

    if not _regulation_is_post_quake(reg, quake_at):
        return {
            "color": "#5b7c99", "dash_array": dash, "show_x": False,
            "weight": 3, "opacity": 0.35 if ended else 0.55,
        }
    if reg.get("content") in FULL_CLOSURE_CONTENTS:
        return {
            "color": "#e60000", "dash_array": dash, "show_x": True,
            "weight": 6, "opacity": 0.5 if ended else 0.95,
        }
    return {
        "color": "#e67e22", "dash_array": dash, "show_x": True,
        "weight": 5, "opacity": 0.5 if ended else 0.9,
    }


@contextmanager
def _deterministic_branca_ids():
    """
    branca(folium)の各要素は既定でランダムなDOM idを振るため、内容が全く同じ
    地図でもPython側で再構築するたびにHTMLが毎回変わってしまい、streamlit_foliumが
    毎クリックごとに地図全体を再マウントしてしまう（＝選択が鈍く感じる主因）。
    要素生成中だけid生成を連番の決定的な値に置き換えて、選択状態が変わらない限り
    生成HTMLが完全に同一になるようにする。
    """
    counter = itertools.count()
    original = branca_element.Element._generate_id

    def _next_id(cls):
        return format(next(counter), "032x")

    branca_element.Element._generate_id = classmethod(_next_id)
    try:
        yield
    finally:
        branca_element.Element._generate_id = original


def point_z_legend_html(point_summary) -> str:
    """
    観測点の色の濃さ（＝異常度）の凡例。地図の外に置く行として返す。

    どちらの地図でも観測点は同じ描き方なので、凡例も同じものを使う。
    以前は転記版のタブの中に直接書いていたため、通れる道マップ版の地図には
    出ていなかった。
    """
    def _sw(style: str) -> str:
        return (
            '<span style="display:inline-block;vertical-align:middle;'
            f'{style}"></span>'
        )

    items = [
        _sw("width:30px;height:11px;border-radius:6px;border:1px solid #aaa;"
            "background:linear-gradient(to right,#deebf7,#08306b);")
        + " 濃く大きいほど異常度|z|が大きい",
    ]
    if int(point_summary["max_abs_z"].isna().sum()):
        items.append(
            _sw("width:11px;height:11px;border-radius:50%;background:#f0f0f0;"
                "border:2px dashed #777;")
            + " 枠が灰色の破線なら地震後が欠測"
        )
    return "".join(
        f'<span style="white-space:nowrap;">{it}</span>' for it in items
    )


class MapSizeFixer(folium.MacroElement):
    """
    隠れたタブの中で作られた地図に、正しい大きさを覚え直させる。

    Leafletは初期化時のコンテナの大きさを覚える。タブが選ばれていない間は
    その中身の大きさが0なので、地図は0×0のつもりのまま動く。すると
      ・背景地図のタイルが1枚しか読み込まれない
      ・クリック位置の判定が画面の半分ほどずれる（canvasで描く図形は
        DOM要素ではないので、この判定でしか拾えず、クリックが効かなくなる）
    という状態になる。実測で、既定でないタブの地図はどれも 0×0 だった。

    大きさが変わったときに invalidateSize を呼んで直す。JSはMacroElementの
    script マクロとして地図の子に付ける。streamlit-folium は地図とその子の
    script マクロだけを集めてJSを組み立てるので、Figureのheader/html/script
    に足したものは実行されない。
    """
    _template = jinja2.Template("""
        {% macro script(this, kwargs) %}
            (function () {
                var m = {{ this._parent.get_name() }};
                var box = m.getContainer();
                var fix = function () {
                    var s = m.getSize();
                    if (s.x !== box.clientWidth || s.y !== box.clientHeight) {
                        m.invalidateSize(false);
                        // canvasで描く図形は、地図が大きさ0のときに作られた
                        // canvas要素も0×0のままで、invalidateSizeだけでは
                        // 描き直されない（タイルは出るのにメッシュが出ない
                        // 状態になる）。描画をやり直させる。
                        m.fire("viewreset");
                    }
                };
                if (window.ResizeObserver) {
                    new ResizeObserver(fix).observe(box);
                }
                window.addEventListener("resize", fix);
                // ResizeObserver に頼らない道も用意する。地図に触れた時点で
                // 直せば、クリックの当たり判定はその後に行われる。
                // pointerdown は capture で拾う（Leafletの判定より先に走らせる。
                // 指で操作する端末では mouseenter が来ない）。
                box.addEventListener("mouseenter", fix);
                box.addEventListener("pointerdown", fix, true);
                setTimeout(fix, 200);
                // ResizeObserver が来ない場合の保険。タブが開かれて大きさが
                // 付いた時点で直り、直ったら止める（最長30秒）。
                var tries = 0;
                var timer = setInterval(function () {
                    tries += 1;
                    fix();
                    var done = box.clientWidth > 0
                        && m.getSize().x === box.clientWidth;
                    if (done || tries > 60) { clearInterval(timer); }
                }, 500);
            })();
        {% endmacro %}
    """)

    def __init__(self):
        super().__init__()
        self._name = "MapSizeFixer"


def add_map_size_fixer(fmap: folium.Map) -> None:
    MapSizeFixer().add_to(fmap)


class MeshHoverLink(folium.MacroElement):
    """
    2枚の地図で、同じメッシュに触れると両方が枠で囲まれるようにする。

    見比べる図なので、片方で見ている場所が反対側のどこなのかが分からないと
    比べられない。メッシュコードで対応づけて、触れているメッシュと同じ
    コードのメッシュを反対側でも強調する。

    DualMapの2枚目（最後に描かれる地図）に付ける。JSは地図の子のscript
    マクロとして出さないと streamlit-folium では実行されない。
    """
    _template = jinja2.Template("""
        {% macro script(this, kwargs) %}
            (function () {
                var maps = [{{ this.other_name }},
                            {{ this._parent.get_name() }}];
                var index = maps.map(function (m) {
                    var byCode = {};
                    m.eachLayer(function (layer) {
                        if (!layer._layers) { return; }
                        Object.keys(layer._layers).forEach(function (key) {
                            var sub = layer._layers[key];
                            var p = sub.feature && sub.feature.properties;
                            if (p && p.mesh) { byCode[p.mesh] = sub; }
                        });
                    });
                    return byCode;
                });
                var HIGHLIGHT = {"weight": 3, "color": "#111111"};
                function paint(sub, on) {
                    if (!sub) { return; }
                    if (on) {
                        sub.setStyle(HIGHLIGHT);
                        if (sub.bringToFront) { sub.bringToFront(); }
                    } else {
                        sub.setStyle({
                            "weight": 0.2,
                            "color": sub.feature.properties.color
                        });
                    }
                }
                index.forEach(function (byCode, i) {
                    var other = index[1 - i];
                    Object.keys(byCode).forEach(function (code) {
                        var sub = byCode[code];
                        sub.on("mouseover", function () {
                            paint(sub, true); paint(other[code], true);
                        });
                        sub.on("mouseout", function () {
                            paint(sub, false); paint(other[code], false);
                        });
                    });
                });
            })();
        {% endmacro %}
    """)

    def __init__(self, other_map: folium.Map):
        super().__init__()
        self._name = "MeshHoverLink"
        # DualMapの2枚の地図はどちらもFigureの子なので、相手を親から辿れない。
        # 変数名を持たせる（streamlit-folium の名前の付け替えも、この文字列に
        # 対して同じように効く）。
        self.other_name = other_map.get_name()


def point_z_legend_row(point_summary) -> str:
    """
    上の凡例を、地図の上の規制の凡例（mlit_map_view.legend_html）の続きの
    行として出す。フォントと行間は legend_html の枠と同じ値にし、枠の
    下マージン分だけ上に詰めて1つの凡例に見せる。
    """
    return (
        '<div style="font-size:0.79rem;line-height:1.45;margin:-4px 0 4px 0;'
        'display:flex;flex-wrap:wrap;gap:1px 12px;align-items:center;">'
        '<b style="white-space:nowrap;">観測点:</b>'
        + point_z_legend_html(point_summary)
        + "</div>"
    )


def point_marker_caption(anomaly_end) -> str:
    """
    観測点マーカーの読み方（濃さ＝異常度・番号の順・震源の記号・選択の仕方）。

    どちらの地図でも観測点は同じ描き方なので、説明も同じものを両タブで使う
    （片方だけ直して文言がずれるのを防ぐ）。
    """
    return (
        "観測点は色が濃いほど地震後の交通量変化（|zスコア|）が大きいことを示す"
        "（青系のグラデーション）。"
        f"**この異常度は発災後1週間（〜{anomaly_end:%m-%d %H:%M}）を"
        "対象に求めた値です**（それ以降はお盆にかかり、地震前の"
        "平常時と比べる意味が薄いため）。"
        "地点番号は異常度の大きい順。青い × は震源"
        "（気象庁の推計震度分布図と同じ記号）。"
        "選択中の観測点は赤/緑の枠で強調表示されます（最大2地点）。"
        "丸いマーカーをクリックしても選択できます（反映に1〜2秒かかります）。"
        "すぐに切り替えたいときは上のプルダウンが速いです。"
    )


POINT_LEGEND_CSS = (
    # 地図内の観測点凡例（右上）。レイヤ一覧(bottomright)と重ならない
    # 位置に置き、クリックは下の地図に通す。
    ".pt-shape-legend{position:absolute;right:8px;top:8px;z-index:650;"
    "background:rgba(255,255,255,0.88);border:1px solid #bbb;border-radius:4px;"
    "padding:4px 7px;font-size:11.5px;line-height:1.6;color:#222;"
    "pointer-events:none;font-family:sans-serif;}"
    ".pt-shape-legend span{display:block;white-space:nowrap;}"
    ".pt-mark{display:inline-block;width:10px;height:10px;margin-right:5px;"
    "background:#4a7ab5;border:1px solid #333;vertical-align:middle;}"
    ".pt-circle{border-radius:50%;}"
    # 角丸の四角。地図のマーカーと同じ丸め方（1辺の28%）にする
    ".pt-square{width:9px;height:9px;border-radius:28%;}"
    # 人口の地図は観測点を白く塗るので、凡例の印も白にする（塗りが違うと
    # 凡例と地図の印が対応しない）。
    ".pt-shape-legend.pt-plain .pt-mark{background:#ffffff;border:2px solid #333;}"
    ".pt-legend-head{font-size:10.5px;color:#555;letter-spacing:0.02em;"
    "border-bottom:1px solid #ddd;margin-bottom:2px;padding-bottom:1px;}"
)


LAYER_CONTROL_CSS = (
    # レイヤ一覧はたたんだ状態で置き、アイコンに載せると開く（Leaflet標準）。
    # 常に開いていると右下の規制の線にかぶって中身が読めないため。
    # たたんだアイコンだけだと何のボタンか分からないので、少し大きめにして
    # 「レイヤ」の字を入れる。
    ".leaflet-control-layers-toggle{width:74px !important;height:26px !important;"
    "background-image:none !important;background:rgba(255,255,255,0.92);"
    "border:1px solid #bbb;border-radius:4px;"
    "font:600 11.5px/26px sans-serif;color:#333;text-align:center;}"
    ".leaflet-control-layers-toggle:after{content:'≡ レイヤ';}"
    ".leaflet-control-layers-expanded .leaflet-control-layers-toggle{display:none;}"
)


def point_legend_html(point_summary: pd.DataFrame,
                      plain: bool = False) -> str:
    """
    観測点の形（＝道路の種別）の凡例。地図を見ながら参照するものなので
    地図の外ではなく地図の中（右上）に置く。

    観測点の描き方はどの地図でも同じなので、通れる道マップ版の地図にも
    同じものを渡す。区分名（一般国道／高速自動車国道）はJARTICが返す
    道路種別をそのままで、道路の法的な位置づけを表すものではない
    （例: 九州中央自動車道は種別1だが、規制を公表しているのは直轄国道の
    管理者である熊本河川国道事務所）。この但し書きはREADMEとdocsにある。
    """
    # 形の意味は地図に何が出ているかで変わらないので、凡例は固定で出す
    # （以前は高速の点が無いと凡例そのものを出しておらず、●が何を指すのか
    # 分からなくなっていた）。
    n_general, n_express, n_unknown = _road_type_counts(point_summary)
    rows = [
        '<span class="pt-legend-head">観測点の道路種別</span>',
        f'<span><i class="pt-mark pt-circle"></i>一般国道 {n_general}点</span>',
        f'<span><i class="pt-mark pt-square"></i>高速自動車国道 {n_express}点</span>',
    ]
    if n_unknown:
        # 形は●のままだが、一般国道と言い切れないので別行にする
        rows.append(
            f'<span><i class="pt-mark pt-circle"></i>種別不明 {n_unknown}点</span>'
        )
    css_class = "pt-shape-legend pt-plain" if plain else "pt-shape-legend"
    return f'<div class="{css_class}">' + "".join(rows) + "</div>"


def _map_center(point_summary: pd.DataFrame, mainshock: dict) -> list:
    """
    地図の中心を、観測点を1点も欠かさない範囲で震源にいちばん近づける。

    観測点の平均で決めると中心が震源の北東にずれ、震源が画面の隅に寄って
    「どこで起きた地震か」が読みにくかった。かといって震源そのものを中心に
    すると、縮尺を変えない以上どちらかの端の観測点が画面から出てしまう。

    そこで縮尺（MAP_ZOOM）と画面の大きさから「この中心なら全点が入る」
    範囲を出し、そこへ震源を丸める。全点が入る中心が無いときだけ、
    観測点の真ん中に戻す。
    """
    lat0, lat1 = point_summary["point_lat"].min(), point_summary["point_lat"].max()
    lon0, lon1 = point_summary["point_lon"].min(), point_summary["point_lon"].max()
    # Webメルカトルの1度あたりピクセル数（緯度は経度をcosで割った値）
    px_per_deg_lon = 256 * (2 ** MAP_ZOOM) / 360
    px_per_deg_lat = px_per_deg_lon / math.cos(math.radians((lat0 + lat1) / 2))
    half_lat = (MAP_HEIGHT_PX / 2 - MAP_EDGE_MARGIN_PX) / px_per_deg_lat
    half_lon = (MAP_MIN_WIDTH_PX / 2 - MAP_EDGE_MARGIN_PX) / px_per_deg_lon

    def _clamp(target, lo_pt, hi_pt, half):
        lo, hi = hi_pt - half, lo_pt + half   # 全点が入る中心の下限・上限
        if lo > hi:                            # 縮尺の都合で入りきらない
            return (lo_pt + hi_pt) / 2
        return min(max(target, lo), hi)

    return [
        _clamp(mainshock["epicenter_lat"], lat0, lat1, half_lat),
        _clamp(mainshock["epicenter_lon"], lon0, lon1, half_lon),
    ]


def build_base_map(
    point_summary: pd.DataFrame,
    mainshock: dict,
    regulations: list = None,
    mlit: dict = None,
    point_labels: dict = None,
    nexco: dict = None,
) -> folium.Map:
    """
    選択状態に依存しない「ベース地図」（背景タイル・通行規制・震源）を作る。

    観測点マーカーは選択状態で見た目が変わるため、この地図には含めず
    st_folium の feature_group_to_add で別途渡す。streamlit_folium が
    コンポーネントの同一性判定に使うハッシュはベース地図のJSだけから
    作られる（feature_group は含まれない）ため、こう分けておくと
    選択のたびに地図全体が再マウント（タイル再読込・全要素再構築）されず、
    観測点レイヤーだけが差し替わる。

    folium.Map.render()は副作用を持ち複数回呼ぶと壊れるため、Mapオブジェクト
    自体はキャッシュしない（毎回作り直す）。

    避難所（748件）は、個別表示だとクリック応答が重く、クラスタ化すると
    個々の場所が分からず意味が薄いため、地図には表示しない。
    """
    with _deterministic_branca_ids():
        fmap = folium.Map(
            location=_map_center(point_summary, mainshock),
            zoom_start=MAP_ZOOM, tiles=None,
        )
        # 通行規制などの重ね合わせ情報が見やすいよう、背景地図は半透明にする。
        # 背景地図は1種類しかなく、切り替えられないものをレイヤ一覧に出しても
        # 幅を取るだけなので control=False で一覧から外す。
        folium.TileLayer(
            "OpenStreetMap", name="OpenStreetMap", opacity=0.55, control=False,
        ).add_to(fmap)
        # st_folium(width=550) は地図divに width:550px を焼き込む。iframe側は
        # CSSで100%にしているため、列が550pxより狭くなると地図だけが
        # iframeからあふれ、右端に貼り付くレイヤ一覧が画面外にはみ出す。
        # 列幅はブラウザの表示倍率で変わる（100%は90%より狭い）ので、
        # 地図div自体をiframe幅に追従させる。
        #
        # あわせてレイヤ一覧の幅にも上限を掛ける。%指定は親が leaflet-top/right の
        # 隅要素になって地図幅に対して解決しないため、iframeのビューポート幅である
        # vw を使う（iframe幅＝地図幅）。折り返しはラベルのテキスト部分
        # （一番内側のspan）だけに許す。labelや外側のspanに white-space:normal を
        # 当てると、チェックボックスとテキストの間で改行され切れて見えてしまう。
        fmap.get_root().header.add_child(folium.Element(
            "<style>"
            "#map_div{width:100% !important;}"
            # 通行止めの×は目印だけなのでクリックを透過させ、下にある
            # 観測点マーカーを確実にクリックできるようにする。
            ".reg-x-mark{pointer-events:none !important;}"
            + POINT_LEGEND_CSS + LAYER_CONTROL_CSS +
            # ツールチップは既定（white-space:nowrap）だと横に伸びて地図の外まで
            # はみ出す。かといって max-width だけを付けると、絶対配置の
            # 幅が「その位置から地図の端までの残り幅」に縮められ、
            # 地図の端に近い規制では数文字ごとに折り返す細長い箱になる。
            # width:max-content で内容なりの幅を確保し、その上で上限を掛ける。
            ".leaflet-tooltip{width:max-content;max-width:260px;"
            "white-space:normal;font-size:11.5px;line-height:1.5;padding:5px 8px;}"
            ".leaflet-control-layers{font-size:12px;max-width:calc(100vw - 28px);}"
            ".leaflet-control-layers-overlays label,"
            ".leaflet-control-layers-overlays label>span{white-space:nowrap;}"
            ".leaflet-control-layers-overlays label>span>span"
            "{white-space:normal;overflow-wrap:anywhere;}"
            "</style>"
        ))

        # ここで fit_bounds は使わない。st_folium は folium 側の location と
        # zoom だけを見て地図を組み立てるため、fit_bounds を書いても効かない
        # （実際、以前の地図は fit_bounds ではなく観測点の平均座標で
        # 中心が決まっていた）。中心と縮尺は _map_center / MAP_ZOOM で決める。

        # ((状態, データソース), レイヤ)。状態は 0=規制中 / 1=解除済み /
        # 2=地震前からの規制、データソースは 0=高速道路 / 1=直轄国道 /
        # 2=県・市町村管理（管理者ごとに公表の経路が違うので、それが区切り）。
        # 県・市町村管理には、県が管理する補助国道（国道219号など）も入る。
        # 規制中を上、解除済みをその下、地震前を一番下に並べ、
        # 同じ状態のなかでは直轄国道を先に置く。
        overlays = []

        if regulations:
            now = _now_jst()
            quake_at = datetime.fromisoformat(mainshock["occurred_at"]).replace(tzinfo=None)
            # レイヤは「どこから来たデータか（道路の管理者）」×「いまどの状態か」
            # で分ける。記録している規制の多くはすでに解除済みなので、
            # 既定では規制中だけを出し、解除済みと地震前からの規制は
            # 地図の右下のレイヤ一覧でオンにしてもらう。
            post_active_layer = folium.FeatureGroup(name="県・市町村管理：規制中")
            post_ended_layer = folium.FeatureGroup(
                name="県・市町村管理：解除済み", show=False
            )
            # 名前は他のレイヤと同じ長さに収める。ここが1つだけ長いと
            # レイヤ一覧の幅がそれに引きずられて、地図の右下で場所を取る。
            # 「工事・過去の災害等」という中身は地図の上の凡例に書いてある。
            pre_layer = folium.FeatureGroup(
                name="県・市町村管理：地震前から", show=False
            )
            for reg in regulations:
                is_post = _regulation_is_post_quake(reg, quake_at)
                style = _regulation_style(reg, now, quake_at)
                ended = _regulation_is_ended(reg, now)
                period = reg["start_timestamp"] or "?"
                period += f" 〜 {reg['end_timestamp']}" if ended and reg["end_timestamp"] else " 〜 (継続中)"
                if not is_post:
                    target = pre_layer
                else:
                    target = post_ended_layer if ended else post_active_layer
                folium.PolyLine(
                    locations=reg["path"],
                    color=style["color"],
                    weight=style["weight"],
                    opacity=style["opacity"],
                    dash_array=style["dash_array"],
                    tooltip=(
                        f"{reg['route_name']}（{reg['region']}）<br>"
                        f"<b>{'今回の地震以降に開始' if is_post else '地震前からの規制'}"
                        f"／{'解除済み' if ended else '規制中'}</b><br>"
                        f"{reg['content']}｜{reg['reason_type']}"
                        f"{('・' + reg['reason_detail']) if reg['reason_detail'] else ''}<br>"
                        f"{period}"
                    ),
                ).add_to(target)
                if style["show_x"]:
                    mid = reg["path"][len(reg["path"]) // 2]
                    folium.Marker(
                        location=mid,
                        # ×はただの目印で、ツールチップもポップアップも持たない。
                        # マーカーはmarkerPane(z600)にあり観測点の円(overlayPane z400)より
                        # 上に来るため、そのままだと観測点クリックを横取りしてしまう
                        # （クリックが×に当たると、その座標が返るだけで観測点が選ばれない）。
                        # クリックを透過させるクラスを付ける。
                        icon=_x_circle_icon(style["color"], ended),
                    ).add_to(target)
            # 直轄国道のレイヤも作ってから、まとめて決まった順に足す
            # （足した順がそのままレイヤ一覧の並びになるため）。
            overlays += [
                ((0, 2), post_active_layer), ((1, 2), post_ended_layer),
                ((2, 2), pre_layer),
            ]

        # 直轄国道の規制。PDFは区間を「○○IC〜○○IC」と名前で示すだけで座標が無いが、
        # 実在する道路区間なので端点のIC座標から線形を復元してある
        # （scripts/build_mlit_paths.py が path として書き込む）。
        # 線形を作れなかった規制（キロポストで示された地点など）は、
        # 掛かっていた観測点が分かる場合だけ▲を置いてフォールバックする。
        mlit_items = (mlit or {}).get("items", [])
        mlit_drawn = 0
        if mlit_items:
            mlit_active_layer = folium.FeatureGroup(name="直轄国道：規制中")
            mlit_ended_layer = folium.FeatureGroup(
                name="直轄国道：解除済み", show=False
            )

            def mlit_layer_for(is_ended: bool):
                return mlit_ended_layer if is_ended else mlit_active_layer
            # 同じ区間に別の日時の規制が複数あることがある（阿蘇西IC〜車帰IC は
            # 本震直後の通行止めと、8月の夜間工事の通行止めの2件）。
            # 区間ごとに1本にまとめないと線が完全に重なり、下になった方は
            # ツールチップを出せず、その規制の存在に気づけない。
            by_section = {}
            for item in mlit_items:
                if not item.get("path"):
                    continue
                by_section.setdefault(
                    (item["route_name"], item["section"]), []
                ).append(item)

            for (route_name, section), items in by_section.items():
                # 見た目は「いま規制中のものがあるか」「全面通行止めが
                # 含まれるか」で決める（重い方に寄せる）。
                ended = all(i.get("end_timestamp") for i in items)
                full = any(i.get("content") in FULL_CLOSURE_CONTENTS for i in items)
                periods = "".join(
                    f"<b>{i['content']}／{'解除済み' if i.get('end_timestamp') else '規制中'}</b><br>"
                    f"{i['start_timestamp']} 〜 {i['end_timestamp'] or '(継続中)'}<br>"
                    for i in sorted(items, key=lambda x: x["start_timestamp"])
                )
                count = f"（{len(items)}件）" if len(items) > 1 else ""
                folium.PolyLine(
                    locations=items[0]["path"],
                    color="#e60000" if full else "#e67e22",
                    weight=6 if full else 5,
                    opacity=0.5 if ended else 0.95,
                    dash_array="6,8" if ended else None,
                    tooltip=(
                        f"<b>{route_name}</b><br>{section}{count}<br>"
                        f"{periods}"
                        f"直轄国道（出典: 熊本河川国道事務所）"
                    ),
                ).add_to(mlit_layer_for(ended))
                # 県フィードの地震後の規制と同じく、線の中点にも⊗を置く。
                # 直轄国道の規制はいずれも今回の地震以降のものなので、
                # 「⊗＝地震後に始まった規制」という凡例の規則に合わせる。
                folium.Marker(
                    location=items[0]["path"][len(items[0]["path"]) // 2],
                    icon=_x_circle_icon("#e60000" if full else "#e67e22", ended),
                ).add_to(mlit_layer_for(ended))
                mlit_drawn += 1

            # 区間の線を引けない規制のうち、場所が1地点として特定できたものは
            # その位置に印を置く（キロポストや地名でしか示されていないもの）。
            # 位置は地名から復元しているので誤差があり、その旨をツールチップに出す。
            for item in mlit_items:
                pt = item.get("point")
                if not pt or item.get("path"):
                    continue
                ended = bool(item.get("end_timestamp"))
                full = item.get("content") in FULL_CLOSURE_CONTENTS
                color = "#e60000" if full else "#e67e22"
                # 観測点マーカー（塗りつぶしの丸）と見分けがつくよう、
                # PDFの別添図と同じ ⊗ で描く。
                folium.Marker(
                    location=[pt["lat"], pt["lon"]],
                    icon=_x_circle_icon(color, ended),
                    tooltip=(
                        f"<b>{item['route_name']}</b><br>{item['section']}<br>"
                        f"<b>{item['content']}／{'解除済み' if ended else '規制中'}</b><br>"
                        f"{item['start_timestamp']} 〜 "
                        f"{item['end_timestamp'] or '(継続中)'}<br>"
                        "位置は地名から復元した概略値<br>"
                        "直轄国道（出典: 熊本河川国道事務所）"
                    ),
                ).add_to(mlit_layer_for(ended))
                mlit_drawn += 1

            # 区間の端点（IC）を点で落とす。線だけだと、どのICからどのICまで
            # なのかが地図から読めないため。同じICが複数の区間の端点になる
            # （車帰ICは大津IC側と阿蘇西IC側の両方）ので、ICごとに1点だけ置き、
            # 属する区間をツールチップに並べる。
            # 観測点マーカーより小さく、白抜きにして混同しないようにする。
            # ICは区間とセットで意味を持つので、その区間と同じレイヤに入れる。
            # 同じICが規制中の区間と解除済みの区間の両方の端点になることが
            # あるため、レイヤごとに1つずつ置く（片方を消してももう片方に残る）。
            ic_points = {}
            for (route_name, section), items in by_section.items():
                layer = mlit_layer_for(
                    all(i.get("end_timestamp") for i in items)
                )
                for ep in items[0].get("endpoints") or []:
                    # 同じICでも報によって座標が数十mずれていることがあり、
                    # 座標で束ねると同じICが2つ描かれる。名前で束ねる。
                    key = (ep["name"], id(layer))
                    entry = ic_points.setdefault(
                        key,
                        {"name": ep["name"], "lat": ep["lat"], "lon": ep["lon"],
                         "node": ep.get("osm_node"), "sections": [],
                         "layer": layer},
                    )
                    label = f"{route_name} {section}"
                    if label not in entry["sections"]:
                        entry["sections"].append(label)
            for info in ic_points.values():
                folium.CircleMarker(
                    location=[info["lat"], info["lon"]],
                    radius=5,
                    color="#333333",
                    weight=2,
                    fill=True,
                    fill_color="#ffffff",
                    fill_opacity=0.95,
                    tooltip=(
                        f"<b>{info['name']}</b><br>"
                        + "".join(f"{sec} の端点<br>" for sec in info["sections"])
                        + "位置はOpenStreetMapのICノード"
                        + (f"（node/{info['node']}）" if info["node"] else "")
                    ),
                ).add_to(info["layer"])

            # 線形が無い規制のフォールバック（掛かっていた観測点の上に▲）
            if not point_summary.empty:
                max_z_for_radius = point_summary["max_abs_z"].max()
                if not (pd.notna(max_z_for_radius) and max_z_for_radius > 0):
                    max_z_for_radius = 1.0
                for _, row in point_summary.iterrows():
                    hits = [
                        h for h in mlit_regulations_for_point(mlit, row.get("point_code"))
                        if not h.get("path")
                    ]
                    if not hits:
                        continue
                    label = (point_labels or {}).get(row["point_id"], row["point_id"])
                    lines = "<br>".join(
                        f"{h['route_name']}（{h['section']}）{h['content']}<br>"
                        f"{h['start_timestamp']} 〜 {h['end_timestamp'] or '(継続中)'}"
                        for h in hits
                    )
                    folium.Marker(
                        location=[row["point_lat"], row["point_lon"]],
                        icon=folium.DivIcon(
                            html=(
                                '<div style="font-size:18px;font-weight:900;color:#b00000;'
                                'line-height:1;text-shadow:0 0 3px white,0 0 3px white;">▲</div>'
                            ),
                            # 観測点の丸に重ねると観測点の属性のように見えるので外側（上）に置く。
                            # 選択時は半径が+3されるためその分も見込む。
                            icon_size=(18, 18),
                            icon_anchor=(9, 18 + _point_radius(row["max_abs_z"], max_z_for_radius) + 6),
                        ),
                        tooltip=(
                            f"<b>直轄国道の規制</b>（区間の線形が作れないため位置のみ）<br>"
                            f"{lines}<br>{label} に掛かっていたもの<br>"
                            "出典: 熊本河川国道事務所（県ポータルのデータには含まれません）"
                        ),
                    ).add_to(mlit_active_layer)
                    mlit_drawn += 1

            if mlit_drawn:
                overlays += [
                    ((0, 1), mlit_active_layer), ((1, 1), mlit_ended_layer),
                ]

        # 高速道路（NEXCO西日本）の規制。すべて区間で示されているので、
        # 直轄国道のような地点・▲のフォールバックは要らない。
        nexco_items = (nexco or {}).get("items", [])
        if nexco_items:
            nx_active = folium.FeatureGroup(name="高速道路：規制中")
            nx_ended = folium.FeatureGroup(name="高速道路：解除済み", show=False)
            nx_ic = {}
            for item in nexco_items:
                path = item.get("path")
                if not path:
                    continue
                ended = bool(item.get("end_timestamp"))
                full = item.get("content") in FULL_CLOSURE_CONTENTS
                color = "#e60000" if full else "#e67e22"
                layer = nx_ended if ended else nx_active
                folium.PolyLine(
                    locations=path,
                    color=color, weight=6 if full else 5,
                    opacity=0.5 if ended else 0.95,
                    dash_array="6,8" if ended else None,
                    tooltip=(
                        f"<b>{item['route_name']}</b><br>{item['section']}<br>"
                        f"<b>{item['content']}／"
                        f"{'解除済み' if ended else '規制中'}</b><br>"
                        f"{item['start_timestamp']} 〜 "
                        f"{item['end_timestamp'] or '(継続中)'}<br>"
                        + _emergency_access_html(item)
                        + "高速道路（出典: NEXCO西日本）"
                    ),
                ).add_to(layer)
                folium.Marker(
                    location=path[len(path) // 2],
                    icon=_x_circle_icon(color, ended),
                ).add_to(layer)
                for ep in item.get("endpoints") or []:
                    # 座標ではなく名前で束ねる（上と同じ理由）
                    key = (ep["name"], id(layer))
                    entry = nx_ic.setdefault(
                        key,
                        {"name": ep["name"], "lat": ep["lat"], "lon": ep["lon"],
                         "node": ep.get("osm_node"), "sections": [],
                         "layer": layer},
                    )
                    label = f"{item['route_name']} {item['section']}"
                    if label not in entry["sections"]:
                        entry["sections"].append(label)
                # 本文で言及されているIC・SA・PA。「川田橋（宇城氷川SIC〜
                # 八代IC）」のように文章にだけ出てきて、どこの話なのかが
                # 地図から追えなかったので、区間の端点と同じ描き方で置く。
                for mp in item.get("mentioned_points") or []:
                    folium.CircleMarker(
                        location=[mp["lat"], mp["lon"]],
                        radius=4, color="#555555", weight=2,
                        fill=True, fill_color="#ffffff", fill_opacity=0.95,
                        tooltip=(
                            f"<b>{mp['name']}</b><br>"
                            f"{item['route_name']}（{item['section']}）の本文で"
                            f"言及<br>{mp['note']}<br>"
                            f"位置はOpenStreetMap（{mp['osm']}）"
                        ),
                    ).add_to(layer)
            for info in nx_ic.values():
                folium.CircleMarker(
                    location=[info["lat"], info["lon"]], radius=5,
                    color="#333333", weight=2,
                    fill=True, fill_color="#ffffff", fill_opacity=0.95,
                    tooltip=(
                        f"<b>{info['name']}</b><br>"
                        + "".join(f"{sec} の端点<br>" for sec in info["sections"])
                        + "位置はOpenStreetMapのICノード"
                        + (f"（node/{info['node']}）" if info["node"] else "")
                    ),
                ).add_to(info["layer"])
            overlays += [((0, 0), nx_active), ((1, 0), nx_ended)]

        if regulations or mlit_drawn:
            # 左上・右上だと観測点マーカーに被るので右下に置く
            for _, layer in sorted(overlays, key=lambda x: x[0]):
                layer.add_to(fmap)
            folium.LayerControl(position="bottomright", collapsed=True).add_to(fmap)

        legend = point_legend_html(point_summary)
        if legend:
            fmap.get_root().html.add_child(folium.Element(legend))

        folium.Marker(
            location=[mainshock["epicenter_lat"], mainshock["epicenter_lon"]],
            icon=_epicenter_icon(),
            popup=folium.Popup(
                f"震源（本震）: {mainshock['epicenter_name']}<br>"
                f"M{mainshock['magnitude']} 最大震度{mainshock['max_intensity']}",
                max_width=250,
            ),
            tooltip=(
                f"震源（本震） M{mainshock['magnitude']}"
                f" 最大震度{mainshock['max_intensity']}"
            ),
        ).add_to(fmap)

        return fmap


REPO_URL = "https://github.com/centrixuge/kumamoto-earthquake-traffic-map"


def render_source_table(regulations: list, n_post: int,
                        mlit: dict, nexco: dict) -> str:
    """
    通行規制をどこから集めているかを表にする。

    管理者ごとに公表の手段が違い、県はJSON配信、国と高速道路はPDFしか
    無い。この事情を文章で説明すると長くなって読まれないので、
    「出典の公表ページ」と「手元に集めたデータ」の2つのリンクを
    行ごとに持つ表にする。
    """
    def _cell(label: str, url: str) -> str:
        return f'<a href="{url}" target="_blank">{label}</a>' if url else label

    rows = [
        (
            f"① 県・市町村管理の道路<br>（補助国道・県道・市町村道）<br>{len(regulations)}件（地震後 {n_post}件）",
            _cell("防災情報くまもと<br>通行規制情報",
                  "https://portal.bousai.pref.kumamoto.jp/?p=traffic"),
            "公開JSONを自動取得<br>"
            + _cell("取得したデータ",
                    f"{REPO_URL}/blob/main/data/archive/regulations_archive.json"),
        ),
        (
            f"② 直轄国道<br>{len((mlit or {}).get('items', []))}件",
            _cell("熊本河川国道事務所<br>事務所からのお知らせ",
                  mlit.get("source_url", "")),
            "PDFを手作業で転記<br>"
            + _cell("収集したPDF",
                    f"{REPO_URL}/tree/main/{mlit.get('pdf_dir', 'data')}"),
        ),
        (
            f"③ 高速道路<br>{len((nexco or {}).get('items', []))}件",
            _cell("NEXCO西日本<br>九州支社", nexco.get("source_url", "")),
            "PDFを手作業で転記<br>"
            + _cell("収集したPDF",
                    f"{REPO_URL}/tree/main/{nexco.get('pdf_dir', 'data')}"),
        ),
    ]
    td = "padding:3px 6px;border-bottom:1px solid #e6e6e6;vertical-align:top;"
    body = "".join(
        "<tr>" + "".join(f'<td style="{td}">{c}</td>' for c in row) + "</tr>"
        for row in rows
    )
    return (
        '<table style="width:100%;border-collapse:collapse;font-size:0.75rem;'
        'line-height:1.4;margin:2px 0 6px 0;">'
        '<thead><tr>'
        + "".join(
            f'<th style="{td}text-align:left;white-space:nowrap;">{h}</th>'
            for h in ("規制の対象", "出典（公表ページ）", "集め方")
        )
        + "</tr></thead><tbody>" + body + "</tbody></table>"
    )


def build_plain_points_feature_group(
    point_summary: pd.DataFrame, point_labels: dict, selected_points=(),
    interactive: bool = True, into: folium.FeatureGroup = None,
) -> folium.FeatureGroup:
    """
    人口の地図に置く観測点。異常度で色や大きさを変えず、小さな同じ印にする。

    人口の地図で見たいのはメッシュの色と選択枠なので、大きく塗られた観測点が
    その上に乗ると肝心のメッシュが読めない。ここでは「観測点がどこにあるか」と
    「いま選んでいるか」だけ分かればよいので、異常度の凡例も出さない
    （異常度で描き分けた地図は交通量のタブにある）。

    形（●＝一般国道／■＝高速自動車国道）は他の地図と揃える。大きさは同じ
    なので地図は混まないし、揃えておかないと同じ観測点が別の記号で出る。
    """
    # into が渡されたらそこに足す（選択の表示と同じ入れ物にまとめるため）
    fg = into if into is not None else folium.FeatureGroup(name="観測点")
    for _, row in point_summary.iterrows():
        is_selected = row["point_id"] in selected_points
        label = point_labels.get(row["point_id"], row["point_id"])
        tooltip = f"{label}<br>{POINT_TOOLTIP_HINT}" if interactive else None
        edge = POINT_CHART_COLOR if is_selected else "#333333"
        weight = 3 if is_selected else 2
        radius = 9 if is_selected else 7

        if is_selected and interactive:
            # 白い縁取りを下に敷いて、下の塗りに関係なく輪郭を出す
            # （当たり判定からは外し、上の印でクリックを拾う）。
            halo = _plain_point_shape(row, radius, "#ffffff", 6, None, halo=True)
            halo.options["interactive"] = False
            halo.add_to(fg)
        marker = _plain_point_shape(row, radius, edge, weight, tooltip)
        if not interactive:
            # メッシュを選んでいる間は、観測点がその上でクリックを拾わない
            marker.options["interactive"] = False
        marker.add_to(fg)
    return fg


def _plain_point_shape(row, radius: int, edge: str, weight: int, tooltip,
                       halo: bool = False):
    """
    人口の地図用の観測点の印。高速自動車国道は四角、それ以外は丸。

    四角はCircleMarkerでは描けないのでDivIconで作る（交通量の地図と同じ作り）。
    ツールチップの文字列は丸と同じにしてあるので、クリックの判定
    （point_id_from_tooltip）は形が変わっても同じように効く。

    halo=True のときは、下に敷く白い座（枠線ではなく一段大きい白い形）を返す。
    四角は枠線が内側に伸びるので、白を外へ出すには形ごと大きくする必要がある。
    """
    location = [row["point_lat"], row["point_lon"]]
    if _is_square_point(row.get("road_type")):
        size = max(6, 2 * radius) + (6 if halo else 0)
        style = (
            f"width:{size}px;height:{size}px;box-sizing:border-box;"
            "border-radius:28%;background:#ffffff;"
            + ("" if halo else f"border:{weight}px solid {edge};")
        )
        return folium.Marker(
            location=location,
            icon=folium.DivIcon(
                icon_size=(size, size), icon_anchor=(size // 2, size // 2),
                html=f'<div style="{style}"></div>',
            ),
            tooltip=tooltip,
        )
    return folium.CircleMarker(
        location=location, radius=radius, color=edge, weight=weight,
        fill=True, fill_color="#ffffff", fill_opacity=1.0, tooltip=tooltip,
    )


def build_points_feature_group(
    point_summary: pd.DataFrame, point_labels: dict, selected_points=()
) -> folium.FeatureGroup:
    """
    観測点マーカーだけを含むFeatureGroupを作る。選択状態で見た目が変わるのは
    このレイヤーだけなので、st_folium の feature_group_to_add に渡すことで
    地図全体の再マウントなしに差し替えられる。
    """
    max_z = point_summary["max_abs_z"].max()
    max_z = max_z if pd.notna(max_z) and max_z > 0 else 1.0

    fg = folium.FeatureGroup(name="観測点")
    for _, row in point_summary.iterrows():
        # 地震後が全て欠測だと zスコアが定義できず max_abs_z が NaN になる。
        # そのまま半径の計算に入れると NaN になってマーカーが描画されず、
        # 観測点が地図から消えてしまう（「異常なし」と区別できない）。
        # 欠測は欠測として、灰色・破線の枠で別に描く。
        no_data = pd.isna(row["max_abs_z"])
        frac = 0.0 if no_data else row["max_abs_z"] / max_z
        is_selected = row["point_id"] in selected_points
        sel_idx = list(selected_points).index(row["point_id"]) if is_selected else None
        border_color = SELECTION_COLORS[sel_idx] if is_selected else ("#777777" if no_data else "#333333")
        label = point_labels.get(row["point_id"], row["point_id"])
        if no_data:
            last_at = row.get("last_observed_at")
            # 地図・異常判定は1時間値ベースなので、時刻も1時間値のものだと明示する
            detail = (
                f"地震後のデータがありません（欠測）<br>"
                f"最後に値があった時刻（1時間値）:<br>{last_at:%Y-%m-%d %H:%M}"
                if pd.notna(last_at) else "地震後のデータがありません（欠測）"
            )
        else:
            detail = (
                f"最大|zスコア|: {row['max_abs_z']:.2f}<br>"
                f"異常件数: {int(row['n_anomaly'])}"
            )
        radius = _point_radius(row["max_abs_z"], max_z, is_selected)
        weight = 4 if is_selected else (2 if no_data else 1)
        fill_color = "#f0f0f0" if no_data else _severity_color(frac)
        fill_opacity = 0.6 if no_data else 0.85
        # 定型句は行を分ける。地点名と続けて1行にすると、その1行が
        # ツールチップの幅を決めてしまい、横に長い吹き出しになる。
        tooltip = f"{label}<br>{POINT_TOOLTIP_HINT}<br>{detail}"

        if _is_square_point(row.get("road_type")):
            # CircleMarkerに角を出す方法はないので、DivIconの四角で描く。
            # ツールチップは丸と同じ文字列にしてあるので、クリックの判定
            # （point_id_from_tooltip）は形が変わっても同じように効く。
            #
            # 見え方の調整を2つ入れている。
            #   1辺を直径の 0.9 倍にする … 同じ差し渡しだと四角は面積が
            #     円の1.27倍あり、同じ異常度でも大きく見えてしまう
            #   角を1辺の28%丸める … 直角のままだと1〜2pxの縁が階段状に
            #     見え、丸との並びで不格好になる。丸めると小さいサイズでも
            #     縁がなめらかに出て、それでいて丸とは十分に見分けられる
            #
            # 1辺は必ず偶数にする。アンカーを size//2 で中心に置いている
            # ので、奇数だと半ピクセルずれて縁がにじむ（丸はSVGなので
            # 影響を受けず、四角だけが甘く見えていた）。
            size = max(6, 2 * int(round(radius * 0.9)))
            style = (
                f"width:{size}px;height:{size}px;box-sizing:border-box;"
                f"border-radius:28%;"
                f"background:{_rgba(fill_color, fill_opacity)};"
                f"border:{weight}px {'dashed' if no_data else 'solid'} {border_color};"
            )
            folium.Marker(
                location=[row["point_lat"], row["point_lon"]],
                icon=folium.DivIcon(
                    icon_size=(size, size),
                    icon_anchor=(size // 2, size // 2),
                    html=f'<div style="{style}"></div>',
                ),
                tooltip=tooltip,
            ).add_to(fg)
        else:
            folium.CircleMarker(
                location=[row["point_lat"], row["point_lon"]],
                radius=radius,
                color=border_color,
                weight=weight,
                dash_array="4,3" if no_data else None,
                fill=True,
                fill_color=fill_color,
                fill_opacity=fill_opacity,
                tooltip=tooltip,
            ).add_to(fg)
    return fg


def build_point_labels(point_summary: pd.DataFrame) -> dict:
    """
    観測点の表示名を作る（セレクタ・地図・グラフ凡例で共用）。
    JARTICの「常時観測点コード」が観測点の公式なIDなので、それをそのまま使う。
    マスタから引けなかった観測点だけ、緯度経度の連番にフォールバックする。
    """
    labels = {}
    for i, (_, row) in enumerate(point_summary.iterrows()):
        code = row.get("point_code")
        labels[row["point_id"]] = (
            f"観測点 {code}" if code and pd.notna(code) else f"地点{i + 1}（コード不明）"
        )
    return labels


def point_id_from_tooltip(tooltip, point_labels: dict):
    """
    クリックされた図形のツールチップ本文から観測点を特定する。

    streamlit_folium の last_object_clicked は「クリックした位置」であって
    マーカーの中心ではない（onLayerClick が event.latlng をそのまま入れている）。
    そのため座標の近さで観測点を当てる方式にすると、
      ・大きなマーカーの端を押すと中心から離れて判定に失敗する
      ・通行規制の線など別の図形を押しただけで近くの観測点が選ばれてしまう
    という取り違えが起きる。ツールチップ本文で判定すれば、
    観測点マーカーを押したときだけ反応し、それ以外は何も起きない。

    ツールチップはHTMLタグを除いたテキストになる（コンポーネント側の
    extractContent が textContent を取る）ので、先頭がラベル＋定型句かで見る。
    改行（<br>）が消えるか空白になるかはコンポーネントの実装次第なので、
    空白を全て落としてから突き合わせる。
    """
    if not tooltip:
        return None
    text = "".join(str(tooltip).split())
    for pid, label in (point_labels or {}).items():
        if text.startswith("".join(f"{label}{POINT_TOOLTIP_HINT}".split())):
            return pid
    return None


def describe_baseline(baseline_windows, daytypes: dict = None) -> str:
    """
    「平常時」が具体的にどの期間を指すのかを説明する文を作る。
    使った日は fetch_and_prepare.py が quake_info.json に書き出した値をそのまま
    受け取るため、説明文と計算内容がずれない。
    """
    if daytypes:
        counts = {dt: len(days) for dt, days in daytypes.items() if days}
        allday = sorted(d for days in daytypes.values() for d in days)
        n = ", ".join(f"{dt} {c}日" for dt, c in counts.items())
        return (
            "**平常時**＝地震発生前の**同じ日区分・同じ時刻**の交通量。"
            "曜日で交通量の形が違うため、月・火・水・木・金はそれぞれ別に平均をとり、"
            "土曜は土曜、日曜と祝日は「日祝」としてまとめています"
            f"（{n}／{allday[0]}〜{allday[-1]}）。"
            "実測の各行は自分と同じ日区分の平常時と比べます。"
            "観測点ごと・日区分ごと・時刻（時）ごとに平均と標準偏差を求めています。"
        )
    # 日区分の情報が無い古いデータ向けのフォールバック
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    days = []
    for w in baseline_windows or []:
        try:
            d = datetime.fromisoformat(w["start"])
        except (ValueError, KeyError, TypeError):
            continue
        days.append((d, f"{d.month}/{d.day}"))
    if not days:
        return "**平常時**＝地震発生前の同じ日区分・同時刻の交通量（時刻帯ごとの平均）。"
    days.sort()
    wd = weekdays[days[0][0].weekday()]
    span = f"{days[0][1]}〜{days[-1][1]}の{wd}曜 {len(days)}日分"
    return (
        f"**平常時**＝地震発生前の同じ曜日（{span}）の、同じ時刻の交通量。"
        "観測点ごと・時刻（時）ごとに平均と標準偏差を求め、今回の実績と比べています。"
    )


def render_timeseries(
    observations: pd.DataFrame, selected_points, quake_at,
    other_event_times=(), point_labels: dict = None, baseline_windows=None,
    unit_label: str = "5分間値", extra_note: str = "",
    mlit_bands=(), baseline_daytypes: dict = None, series_mode: str = "total",
    x_range=None, key_prefix: str = "main",
) -> None:
    """
    key_prefix は、同じ図を複数のタブで出すときに分けるための接頭辞。
    Streamlitは要素のidを中身から自動生成するため、2つの地図のタブで
    2つのタブで同じ設定・同じデータの図を出すとidがぶつかり、
    StreamlitDuplicateElementId で落ちる（新しめのStreamlitで顕在化する）。
    """
    if not selected_points:
        st.info("上のプルダウンから選ぶか、地図上の丸いマーカーをクリックして観測点を選ぶと、ここに交通量の時系列変化が表示されます（最大2地点まで比較可）。")
        return
    if observations.empty:
        st.warning("このビューのデータがまだ生成されていません。`python fetch_and_prepare.py` を実行してください。")
        return

    point_labels = point_labels or {}
    # 地震前日の深夜〜早朝は情報が薄いので、既定の表示は7/27 12:00から始める
    # （データ自体はTARGET_START=7/27 03:00から保持している）。
    # 呼び出し側が期間を指定していればそちらを使う。
    if x_range is None:
        x_range = [TIMESERIES_DISPLAY_START, observations["datetime"].max()]
    else:
        x_range = list(x_range)

    tick_vals, tick_text = _day_ticks(x_range)

    series = SERIES_VEHICLE if series_mode == "vehicle" else SERIES_TOTAL
    chart_height = 261
    for label, sub_series in series:
        missing = [
            suf for suf, _, _ in sub_series
            if f"traffic_{suf}" not in observations.columns
            or f"baseline_mean_{suf}" not in observations.columns
        ]
        if missing:
            st.info(
                f"「{label}」に必要な列がデータにありません（{', '.join(missing)}）。"
                "`python fetch_and_prepare.py` を実行してデータを作り直してください。"
            )
            continue
        fig = go.Figure()
        for i, pid in enumerate(selected_points):
            mark = point_labels.get(pid, pid)
            pdf = observations[observations["point_id"] == pid].sort_values("datetime")
            for suffix, veh_label, color_kind in sub_series:
                palette = SELECTION_ALT_COLORS if color_kind == "alt" else SELECTION_COLORS
                color = palette[i % len(palette)]
                mean_col = f"baseline_mean_{suffix}"
                std_col = f"baseline_std_{suffix}"
                # ±σの帯は1地点のときだけ。車種別では車種ごとに1本ずつ出す
                if len(selected_points) == 1:
                    fig.add_trace(go.Scatter(
                        x=pdf["datetime"], y=pdf[mean_col] + pdf[std_col],
                        mode="lines", line=dict(width=0), showlegend=False,
                    ))
                    fig.add_trace(go.Scatter(
                        x=pdf["datetime"], y=pdf[mean_col] - pdf[std_col],
                        mode="lines", line=dict(width=0), fill="tonexty",
                        fillcolor="rgba(100,100,100,0.2)",
                        name=f"平常時±σ{veh_label}",
                    ))
                fig.add_trace(go.Scatter(
                    x=pdf["datetime"], y=pdf[mean_col],
                    mode="lines", line=dict(color=color, dash="dot", width=1),
                    opacity=0.6, name=f"{mark}{veh_label} 平常時",
                ))
                fig.add_trace(go.Scatter(
                    x=pdf["datetime"], y=pdf[f"traffic_{suffix}"],
                    mode="lines+markers",
                    line=dict(color=color, width=1.2),
                    marker=dict(size=3),
                    name=f"{mark}{veh_label} 実績",
                ))
        # 直轄国道の通行止め期間。この規制は県ポータルのデータに含まれず地図に
        # 線として出ないため、交通量が0になっている理由が図から読めなくなる。
        # 該当観測点を選んだときだけ、期間を帯で示す。
        for band in mlit_bands:
            # 帯は灰色にする。選択中の観測点は赤/緑で描いているので、帯を赤系にすると
            # 緑で選ばれた地点の通行止めが赤帯の上に乗って読みづらくなる。
            fig.add_vrect(
                x0=band["start"], x1=band["end"],
                fillcolor="rgba(110,115,120,0.16)", line_width=0, layer="below",
            )
            fig.add_annotation(
                x=band["start"], y=0.97, yref="paper", xanchor="left", yanchor="top",
                text=band["label"], showarrow=False, align="left",
                font=dict(size=9, color="#4a4f55"),
                bgcolor="rgba(255,255,255,0.75)",
            )
        for t in other_event_times:
            fig.add_vline(x=t, line_dash="dot", line_color="lightgray", line_width=1, opacity=0.7)
        fig.add_vline(x=quake_at, line_dash="dot", line_color="black", line_width=2)
        fig.add_annotation(
            x=quake_at, y=1, yref="paper", yanchor="bottom",
            text="地震発生 16:27", showarrow=False,
            font=dict(size=11, color="black"),
        )
        # 上下2図を並べると縦に長くなり、観測点を比べるたびにスクロールが
        # 必要になるため1図の高さを抑える（400 -> 290 -> 261）。凡例は2図で
        # 同じ内容なので上の図だけに出す。
        #
        # 下余白は2図で同じにする。以前は凡例のない下の図だけ詰めていたが、
        # 図の高さは同じなので下の図の描画領域だけが広くなり（182px と
        # 240px）、上の図が小さく見えていた。凡例が3行になる組み合わせ
        # （2地点×車種別＝12項目）では余白を1行分足す。plotlyの自動拡張に
        # 任せると凡例のある上の図だけ広がってしまうので、行数は項目数から
        # 決めて両方に同じ値を入れる。下の図の下には空きが出るが、
        # 2図は見比べるものなので大きさが揃うほうを取る。
        show_legend = label == series[0][0]
        # 凡例の項目数＝地点数 × その図の系列数 ×（実測・平常時）
        legend_rows = math.ceil(
            len(selected_points) * len(sub_series) * 2 / CHART_LEGEND_PER_ROW
        )
        bottom_margin = CHART_BOTTOM_MARGIN + CHART_LEGEND_ROW_PX * max(
            0, legend_rows - 2
        )
        fig.update_layout(
            height=chart_height,
            # 凡例はグラフ下に置く。上部だとplotlyのモードバー（カメラ・ズーム等の
            # アイコン）と重なり、幅の狭いモバイルでは折り返して読めなくなるため。
            margin=dict(l=10, r=10, t=26, b=bottom_margin),
            showlegend=show_legend,
            # 位置は図の枠（container）を基準にして下端に貼る。描画領域を
            # 基準にすると、描画領域が縮むほど凡例がその下へ回り込み、
            # plotlyが下余白を自動で広げてしまう（62 -> 82）。そうなると
            # 凡例のある上の図だけ描画領域が狭くなる。
            legend=dict(
                orientation="h", yref="container", yanchor="bottom", y=0.012,
                xanchor="left", x=0, font=dict(size=10),
            ),
            xaxis=dict(
                range=x_range,
                tickvals=tick_vals, ticktext=tick_text,
                # 日付の目盛りだけだと日中の位置が読めないので、
                # 6時間ごとの補助線をラベルなしで入れる。
                minor=dict(
                    dtick=6 * 3600 * 1000, showgrid=True,
                    gridcolor="rgba(128,128,128,0.18)",
                ),
            ),
            # 交通量は負にならないので、0より下に軸も罫線も出さない。
            # 平常時±σの帯は下側が負になることがあり（平常時の平均が
            # 小さい深夜帯など）、既定のautorangeだとその分だけ軸が
            # 0より下まで伸びて意味のない目盛りが引かれていた。
            # nonnegative は表示範囲を切るだけで、帯の値自体は変えない。
            yaxis=dict(rangemode="nonnegative"),
        )
        title = f"{label}交通量（{unit_label}）"
        if series_mode == "vehicle":
            title += "　小型＝濃い色／大型＝明るい色"
        st.markdown(f"**{title}**")
        st.plotly_chart(
            fig, use_container_width=True, key=f"ts_{key_prefix}_{label}",
        )

    st.markdown(describe_baseline(baseline_windows, baseline_daytypes))
    st.caption(
        "実線「実績」＝今回の観測実績（地震前日からの推移を含む）。"
        "点線「平常時」＝上記の平常時平均。"
        "灰色の帯「平常時±σ」（1地点選択時のみ）＝平常時の平均±標準偏差で、"
        "この帯から外れているほど平常時と違う動きをしていることを示します。"
        "黒い点線が本震の発生時刻（16:27）、薄いグレーの細い点線がその他の主要な地震"
        f"（震度5弱以上、{len(other_event_times)}件）の発生時刻。"
        + (f" {extra_note}" if extra_note else "")
    )


def tab_fragment(func):
    """
    タブの中身を「そのタブだけ再実行する単位」にする。

    Streamlitは操作のたびにスクリプト全体を回すので、そのままだと1回の
    クリックで5つのタブの地図を作り直す（メッシュ3,923件＋時点比較の2枚ぶん
    7,846件＋規制の地図2枚）。サーバ側は1枚100〜200msでも、その分のHTMLを
    毎回ブラウザへ送ってLeafletに描き直させるので、操作が重くなる。

    fragment にすると、その中の操作ではその中だけが再実行される。
    古いStreamlitには無いので、無ければ素通しにする。
    """
    deco = getattr(st, "fragment", None) or getattr(st, "experimental_fragment", None)
    return deco(func) if deco else func


def rerun_fragment() -> None:
    """fragment の中だけを再実行する（対応していない版では全体を再実行）。"""
    try:
        st.rerun(scope="fragment")
    except TypeError:
        st.rerun()


@tab_fragment
def render_mlit_beta_tab(point_summary: pd.DataFrame, point_labels: dict,
                         quake_at, other_event_times, quake_info: dict,
                         mainshock: dict, anomaly_end) -> None:
    """
    通れる道マップの配布データで規制を描く、既定のタブ。地図と交通量の
    時系列変化を並べて見て、規制と交通量の変化を突き合わせるための画面なので、
    観測点の選択・集計方法・表示期間はそのまま用意する。

    ウィジェットのキーと選択状態は(旧版)タブと別に持つ（同じ画面に2つの
    タブが同時に描かれるため、同じキーは使えない。片方で選んだ地点がもう
    片方に移らないほうが、見比べるときに都合もよい）。
    """
    data = mlit_map_view.load_regulations()
    if data is None:
        st.info(
            "「通れる道マップ」版のデータがありません。"
            "`python scripts/build_mlit_map_regulations.py` で作成してください。"
        )
        return

    st.caption(
        "地図の通行規制は"
        f"[{data['source_name']}]({data['source_url']})の配布データで作っています。"
        "県の公開JSONとPDFからの転記で集めた版は「(旧版)道路規制×交通量」"
        "タブにあり、規制の収録範囲が違うので同じ観測点でも見え方が変わります。  \n"
        f"**規制情報は {data['latest_regulation_time']} 時点のものです**"
        # 配布の回と、その回に入っている規制情報の時点がずれることがある
        # （8/7 16:00 の回は規制情報だけ10:00時点だった）。ずれている
        # ときだけ断りを入れる。
        + (
            f"（配布は {data['latest_snapshot']} の回ですが、そこに入っている"
            "道路規制情報はこの時点のものです）"
            if data["latest_snapshot"] != data["latest_regulation_time"] else ""
        )
        + "。交通量は最新まで出ますが、"
        "この時点より後の規制の変化は地図に反映されていません。",
        unsafe_allow_html=False,
    )

    # 既定の観測点は本家と同じにする（開いた直後に時系列が空だと、
    # 何を見る画面なのかが伝わらないため）
    if "selected_points_beta" not in st.session_state:
        st.session_state["selected_points_beta"] = _default_selection(point_summary)
    selected_points = st.session_state["selected_points_beta"]
    sel_version = st.session_state.get("_sel_version_beta", 0)
    coords_by_id = point_summary.set_index("point_id")[
        ["point_lat", "point_lon"]
    ].to_dict("index")

    def _format_point(pid: str) -> str:
        c = coords_by_id[pid]
        return f"{point_labels[pid]}（{c['point_lat']:.3f}N, {c['point_lon']:.3f}E）"

    # 本家と同じく、地図より前にウィジェットを作る（地図クリックの
    # rerun_fragment() で中断されると、その実行で作られなかった分の状態が落ちる）
    col_select, col_view, col_range, col_clear = st.columns([3, 2, 1.3, 1])
    with col_select:
        picked = st.multiselect(
            f"観測点の選び方：このプルダウンから選ぶか、地図上の丸いマーカーをクリック"
            f"（最大{MAX_SELECTED_POINTS}地点まで並べて比較できます）",
            options=list(point_labels.keys()),
            default=[p for p in selected_points if p in point_labels],
            max_selections=MAX_SELECTED_POINTS,
            format_func=_format_point,
            key=f"point_select_beta_{sel_version}",
        )
    with col_view:
        view_names = list(TIMESERIES_VIEWS.keys())
        saved_view = st.session_state.get("_ts_view_beta", view_names[0])
        view = st.radio(
            "交通量の集計方法", view_names,
            index=view_names.index(saved_view) if saved_view in view_names else 0,
            horizontal=True, key="timeseries_view_beta",
        )
        st.session_state["_ts_view_beta"] = view
        cfg = TIMESERIES_VIEWS[view]
    with col_range:
        range_names = list(TIMESERIES_RANGES.keys())
        saved_range = st.session_state.get("_ts_range_beta", TIMESERIES_DEFAULT_RANGE)
        range_name = st.selectbox(
            "交通量のグラフの表示期間", range_names,
            index=(
                range_names.index(saved_range) if saved_range in range_names
                else range_names.index(TIMESERIES_DEFAULT_RANGE)
            ),
            key="timeseries_range_beta",
        )
        st.session_state["_ts_range_beta"] = range_name
    with col_clear:
        st.write("")
        if st.button("選択をクリア", disabled=not selected_points, key="clear_beta"):
            st.session_state["selected_points_beta"] = []
            st.session_state["_sel_version_beta"] = sel_version + 1
            rerun_fragment()

    selected_points = picked
    st.session_state["selected_points_beta"] = picked

    col_map, col_ts = st.columns([2, 3], gap="large")
    with col_ts:
        st.subheader("選択観測点の交通量の時系列変化（平常時 vs 観測実績）")
    with col_map:
        st.subheader("常時観測点別の異常度 × 通行規制")
        # 地図の上は色分けだけにする。件数の表は縦に場所を取って地図を
        # 押し下げるので、地図の下に置く。
        st.markdown(
            mlit_map_view.legend_html(data)
            + point_z_legend_row(point_summary),
            unsafe_allow_html=True,
        )
        base_map = mlit_map_view.build_map(
            data,
            center=_map_center(point_summary, mainshock),
            zoom=MAP_ZOOM,
            epicenter_icon=_epicenter_icon(),
            epicenter=[mainshock["epicenter_lat"], mainshock["epicenter_lon"]],
            # 観測点の描き方は本家と同じなので、その凡例も同じものを出す
            point_legend=point_legend_html(point_summary),
            point_legend_css=POINT_LEGEND_CSS + LAYER_CONTROL_CSS,
        )
        add_map_size_fixer(base_map)
        points_fg = build_points_feature_group(
            point_summary, point_labels, selected_points
        )
        map_state = st_folium(
            base_map, height=MAP_HEIGHT_PX, width=550,
            feature_group_to_add=points_fg,
            returned_objects=[
                "last_object_clicked_tooltip", "last_object_clicked_count",
            ],
            key="mlit_beta_map_v1",
        )
        st.caption(
            "既定では規制中だけを出しています（地図の右下の「≡ レイヤ」に"
            "マウスを載せると一覧が開きます。一覧の「高速」「国道」"
            "「県・市町村道」は上の3段階の略記）。"
        )
        # 観測点の読み方は(旧版)タブと同じ文言を出す
        st.caption(point_marker_caption(anomaly_end))
        st.caption(
            mlit_map_view.content_note(data)
        )
        note = mlit_map_view.unknown_level_note(data)
        if note:
            st.caption(note)

    # クリックされたのが観測点マーカーのときだけ選択を切り替える
    click_count = map_state.get("last_object_clicked_count") if map_state else None
    if click_count is not None and click_count != st.session_state.get(
        "_last_click_count_beta"
    ):
        st.session_state["_last_click_count_beta"] = click_count
        pid = point_id_from_tooltip(
            map_state.get("last_object_clicked_tooltip"), point_labels
        )
        if pid is not None:
            selected = list(st.session_state["selected_points_beta"])
            if pid in selected:
                selected.remove(pid)
            else:
                if len(selected) >= MAX_SELECTED_POINTS:
                    selected.pop(0)
                selected.append(pid)
            st.session_state["selected_points_beta"] = selected
            st.session_state["_sel_version_beta"] = sel_version + 1
            rerun_fragment()

    with col_ts:
        ts_obs = load_observations(cfg["file"])
        last_at = (
            ts_obs["datetime"].max() if not ts_obs.empty
            else quake_at + pd.Timedelta(days=7)
        )
        render_timeseries(
            ts_obs, selected_points, quake_at, other_event_times, point_labels,
            quake_info.get("hourly_baseline_windows"),
            unit_label=cfg["unit_label"],
            extra_note=cfg["note"],
            # 規制の帯は出さない。この配布データは観測点との対応づけを
            # 持っておらず、近さだけで結びつけると誤った因果を誘発する。
            mlit_bands=(),
            baseline_daytypes=quake_info.get("hourly_baseline_daytypes"),
            series_mode=cfg.get("series", "total"),
            x_range=TIMESERIES_RANGES[range_name](quake_at, last_at),
            key_prefix="beta",
        )

    with st.expander(f"規制の一覧（{len(data['items'])}件）", expanded=False):
        # 表やグラフにも明示のキーを振る。Streamlitは要素のidを中身から
        # 自動生成するため、他のタブと同じ中身になるとidがぶつかる。
        st.dataframe(
            mlit_map_view.regulation_table(data),
            use_container_width=True, height=320, key="mlit_map_table",
        )
        col_csv, col_geo = st.columns(2)
        with col_csv:
            st.download_button(
                f"CSVダウンロード（{len(data['items'])}件）",
                mlit_map_view.regulation_csv(data),
                file_name=mlit_map_view.csv_file_name(data),
                mime="text/csv",
                key="mlit_map_csv",
            )
        with col_geo:
            st.download_button(
                f"GeoJSONダウンロード（{len(data['items'])}件）",
                mlit_map_view.regulation_geojson(data),
                file_name=mlit_map_view.geojson_file_name(data),
                mime="application/geo+json",
                key="mlit_map_geojson",
            )
        st.caption(
            "どちらにも、画面の表に出していない識別子（id）・市町村・規制種別・"
            "初出時点を入れています。**区間の線形が入るのはGeoJSONだけ**です"
            "（CSVの1セルには収まらないため）。GeoJSONは配布元のものと違い、"
            "時点をまたいで突き合わせた結果なので、1件ごとに状態"
            "（規制中／解除済み）と初出・最終確認の時点まで入ります。"
        )
        st.caption(mlit_map_view.source_note(data))


# メッシュの表示範囲。データが出る時間数（336時点中）で絞る。
# 10人未満のメッシュは配信されないため、山間部などは時間帯によって出たり
# 出なかったりする。全部描くと2万件を超えて地図が重くなるので、
# 既定は「常時10人以上」にしておく。
@st.cache_data
def load_holiday_set() -> set:
    """祝日の一覧（日付文字列）。平常時の平均を取る日区分の判定に使う。"""
    path = os.path.join(DATA_DIR, "holidays.json")
    if not os.path.exists(path):
        return set()
    with open(path, encoding="utf-8") as f:
        return set(json.load(f))


# 地図に出すメッシュは「全時点で配信されたもの」に固定する。10人未満は
# 配信されないので、時間帯によって出たり消えたりするメッシュを混ぜると
# 比が同じ土台で比べられないうえ、全部（2万件超）を描くと地図が重い。
MESH_MIN_HOURS = 336
# 比較のタブは地図が主役なので、時系列図と並べる地図より縦を取る。
MESH_COMPARE_MAP_HEIGHT_PX = 620
# メッシュは1つだけ選べるようにする。2つ選べるようにしていたが、居住地の
# 内訳（3段）と交通量まで重なると、どの棒がどれか追えない図になった。
MAX_SELECTED_MESHES = 1
# この地図で選べる観測点は1点だけ。メッシュ2つ＋観測点で、実績と平常時を
# 出すと図が6本になる。これ以上増やすとどの線がどれか追えなくなる。
MAX_SELECTED_POINTS_ON_MESH_MAP = 1
# 初めて開いたときに選んでおくもの（何も選ばれていない図だと、この画面で
# 何が見られるのか伝わらない）。メッシュは宇城市の一角で、夜間人口が
# 発災後に6割まで落ちている場所。
DEFAULT_MESH_ON_MAP = 483075254
DEFAULT_POINT_CODE_ON_MAP = "9110040"
# 地図で「いま選んでいる」ことを示す枠。白の太線を下に敷いて、その上に
# 図と同じ色の線を重ねる（メッシュは青緑、観測点は朱色）。下の白線は
# 塗りの色（青〜赤／黄〜青）に関係なく輪郭を出すためのもの。
SELECTION_HALO_COLOR = "#ffffff"
# 図の色。人口（棒と平常時の階段線）は青緑、交通量は青緑の上でも沈まない
# 赤寄りのオレンジ（朱色）。地図の色分けとは無関係になったので、
# 図の中での見やすさだけで選べる。
MESH_CHART_COLORS = ["#00695C"]
MESH_BAR_COLORS = [("#26A69A", "#9CD8D0", "#E0F2F0")]
POINT_CHART_COLOR = "#DE3F1B"
# 選択中の観測点は、白の太線を下に敷いたうえで枠を図と同じ朱色にする
# （塗りは未選択と同じ白）。

# 人口の地図で「クリックで何を選ぶか」。観測点はメッシュの上に乗るので、
# 両方を同時に押せる状態だと、メッシュを狙っても観測点が拾ってしまう。
# 選んでいないほうは当たり判定から外す（ツールチップも出さない）。
MESH_CLICK_TARGETS = ("メッシュ", "観測点")


def _stale_module_hint(err: Exception) -> str:
    """
    公開側で app.py だけが読み直され、modules/mesh_population.py が古いまま
    動くことが実際に起きている。その形の失敗をここで見分ける。

    以前は「このタブが使う名前」を手で並べて確かめていたが、機能を足すたびに
    並べ直しが要り、実際に追いつかなくなった（VALUE_MODES を足したとき）。
    出た例外そのものから見分ければ、並べ直しは要らない。

    古いモジュールに対して起きるのは次の2つ。
      ・新しい名前を呼ぶ    → AttributeError（module … has no attribute …）
      ・引数が増えた関数を呼ぶ → TypeError（unexpected keyword argument …）
    """
    text = str(err)
    module_name = mesh_population.__name__
    stale = (
        (isinstance(err, AttributeError) and module_name in text)
        or (isinstance(err, TypeError)
            and ("unexpected keyword argument" in text
                 or "positional argument" in text))
    )
    if not stale:
        return ""
    return (
        "\n\n`modules/mesh_population.py` が古いまま読み込まれている状態です"
        "（画面のコードだけが新しく、モジュールが入れ替わっていない）。"
        "**アプリを再起動（Manage app → Reboot）すると直ります。**"
    )


@tab_fragment
def render_mesh_population_tab(point_summary: pd.DataFrame, point_labels: dict,
                               quake_at, mainshock: dict,
                               other_event_times=()) -> None:
    """
    モバイル空間統計の500mメッシュ人口と、常時観測点の交通量を重ねて見るタブ。

    地図にはメッシュと観測点の両方を置き、どちらもクリックで選べる
    （メッシュ2つ・観測点2点まで）。選んだものを1つの図に重ねて出し、
    人がどれだけ残っている場所で交通量がどう動いたかを突き合わせられる
    ようにしている。単位が違うので縦軸は左（人）と右（台/時）に分ける。

    元データは公開できないため、公開repoにはファイルを置かず、集計済みの
    ファイルを非公開の置き場から読む（modules/mesh_population.py）。

    この中の失敗は、このタブの中だけで止める。ここで例外を上げると、
    交通量のタブごとページ全体が落ちてしまうため。
    """
    try:
        _mesh_population_body(
            point_summary, point_labels, quake_at, mainshock, other_event_times
        )
    except mesh_population.MeshDataUnavailable as e:
        st.error(f"モバイル空間統計の集計データを読めませんでした。\n\n{e}")
    except Exception as e:
        st.error(
            f"人口のタブを描けませんでした（{type(e).__name__}: {e}）。"
            "ほかのタブは通常どおり動きます。"
            + _stale_module_hint(e)
        )


def _mesh_population_body(point_summary: pd.DataFrame, point_labels: dict,
                          quake_at, mainshock: dict,
                          other_event_times=()) -> None:
    if not mesh_population.available():
        st.info(
            "モバイル空間統計の集計データが見つかりません。"
            "`python scripts/build_mesh_population.py` で作るか、"
            "非公開の置き場を `st.secrets` に設定してください。"
        )
        return

    meta = mesh_population.load_meta()
    # 列が足りなければ取り直す（集計を作り直した直後に、古い
    # キャッシュを掴んだまま新しい列を読んで落ちるのを防ぐ）
    summary = mesh_population.summary()
    holidays = load_holiday_set()
    start = pd.Timestamp(meta["start"])
    end = pd.Timestamp(meta["end"])

    st.caption(
        f"{meta['source']}（{meta['area']}・{meta['unit']}）。"
        f"収録は **{start:%Y-%m-%d %H:%M} 〜 {end:%Y-%m-%d %H:%M}**"
        f"（{meta['hours']}時点、{meta['meshes']:,}メッシュ）。"
        f"{meta['suppressed_note']}。"
    )

    # 既定の選択。観測点のIDは緯度経度の文字列なので、表示名（観測点コード）
    # から引く。
    default_point = next(
        (pid for pid, label in point_labels.items()
         if label.endswith(DEFAULT_POINT_CODE_ON_MAP)),
        None,
    )
    st.session_state.setdefault("selected_meshes", [DEFAULT_MESH_ON_MAP])
    st.session_state.setdefault(
        "selected_points_mesh", [default_point] if default_point else []
    )
    sel_version = st.session_state.get("_sel_version_mesh", 0)
    selected = list(st.session_state["selected_meshes"])
    selected_points = [
        p for p in st.session_state["selected_points_mesh"]
        if p in set(point_summary["point_id"])
    ]

    col_map, col_ts = st.columns([2, 3], gap="large")
    with col_ts:
        # どのメッシュ・どの観測点かは見出しに出す（系列名に入れると凡例が
        # 長くなり、変なところで折り返す）。
        parts = []
        if selected:
            parts.append(
                f"メッシュ {mesh_population.mesh_label(summary, selected[0])}"
                "の人口推計値"
            )
        if selected_points:
            label = point_labels.get(selected_points[0], selected_points[0])
            parts.append(f"{label} の交通量")
        st.subheader(
            " × ".join(parts) if parts
            else "選択メッシュの人口推計値 × 選択観測点の交通量"
        )
    with col_map:
        st.subheader("500mメッシュ別の推計人口")
        # 操作するものは地図の真上に置く。画面の幅が狭いノートPCでは、
        # 上に離して置くと地図と時系列図を並べたまま切り替えられない。
        # 地図より前に作る点は変えない（地図のクリックで st.rerun したときに、
        # まだ作られていないウィジェットの状態が捨てられるため）。
        pre_days = meta.get("phase_days", {}).get("pre", {})
        post_days = meta.get("phase_days", {}).get("post", {})
        col_hour, col_day = st.columns([1.1, 1])
        with col_hour:
            hours_choices = list(mesh_population.HOUR_METRICS)
            hour_label = st.selectbox(
                "時間帯区分（夜間・昼間・全日）", hours_choices,
                index=hours_choices.index(mesh_population.DEFAULT_HOUR),
                key="mesh_metric",
                help="夜間は3時、昼間は14時を代表時刻にしています"
                     "（どちらも移動の途中が混じりにくい時刻）。"
                     "国勢調査の夜間人口・昼間人口は常住地・従業地で数える"
                     "別の定義なので、その語は使っていません。",
            )
        with col_day:
            day_choices = list(mesh_population.DAY_METRICS)
            day_label = st.selectbox(
                "平日・休日区分", day_choices,
                index=day_choices.index(mesh_population.DEFAULT_DAY),
                key="mesh_daytype",
                help="平日と休日では人の居場所が元から違うので、発災前の"
                     "平日平均には発災後の平日を、休日平均には休日を当てて"
                     "比べます。"
                     f"発災前は平日{pre_days.get('平日', '—')}日・"
                     f"休日{pre_days.get('休日', '—')}日、発災後は平日"
                     f"{post_days.get('平日', '—')}日・"
                     f"休日{post_days.get('休日', '—')}日しかありません。",
            )
        col_res, col_mode = st.columns([1.1, 1])
        with col_res:
            res_choices = list(mesh_population.RESIDENCE_METRICS)
            residence_label = st.selectbox(
                "居住者・来訪者区分", res_choices,
                index=res_choices.index(mesh_population.DEFAULT_RESIDENCE),
                key="mesh_residence",
                help="そのメッシュのある市区町村に住んでいる人か、それ以外か。"
                     "居住者が減って来訪者が増えていれば、避難や支援で人が"
                     "入れ替わったことが読めます。"
                     "「すべて」は総数の配信で、居住者＋来訪者とは一致しません"
                     "（居住地ごとに10人未満だと配信されないため、内訳の合計の"
                     "ほうが小さくなります）。",
            )
        with col_mode:
            mode_choices = list(mesh_population.VALUE_MODES)
            value_mode = st.selectbox(
                "凡例を表示する値", mode_choices,
                index=mode_choices.index(mesh_population.DEFAULT_VALUE_MODE),
                key="mesh_value_mode",
                help="増減率だけだと、元から人が少ない場所の1〜2割の増減が"
                     "都心部と同じ色になります。人口そのものを選ぶと、"
                     "どこに何人いるかで塗り分けます。",
            )
        click_target = st.radio(
            "クリックで選ぶもの", MESH_CLICK_TARGETS, horizontal=True,
            key="mesh_target",
            help="観測点はメッシュの上に乗るので、両方を同時に押せる"
                 "状態だとメッシュを狙っても観測点が拾ってしまいます。"
                 "選んでいないほうは当たり判定から外し、"
                 "カーソルを合わせても何も出ないようにしています。",
        )
        mesh_clickable = click_target == MESH_CLICK_TARGETS[0]
        metric = mesh_population.metric_label(
            hour_label, day_label, residence_label
        )
        geojson = mesh_population.mesh_geojson(
            summary, hour_label, day_label, MESH_MIN_HOURS, value_mode,
            residence_label,
        )
        # 観測点は同じ小さな印にしてあるので、異常度の凡例は出さない
        st.markdown(
            mesh_population.legend_html(metric, value_mode),
            unsafe_allow_html=True,
        )
        fmap = folium.Map(
            # 県全体は1画面に入らないので、震源を中心に置いて熊本市・益城の
            # あたりから見てもらう（縮尺は他の地図と揃える）
            location=[mainshock["epicenter_lat"], mainshock["epicenter_lon"]],
            # 500mメッシュは他の地図の縮尺だと小さすぎて押しにくいので、
            # 1段階寄せた状態から始める
            zoom_start=MAP_ZOOM + 1, tiles=None,
            # 数千のポリゴンを個別のDOM要素にするとブラウザが持たないので、
            # canvasにまとめて描く
            prefer_canvas=True,
        )
        folium.TileLayer(
            "OpenStreetMap", name="OpenStreetMap", opacity=0.45, control=False,
        ).add_to(fmap)
        fmap.get_root().header.add_child(folium.Element(
            "<style>#map_div{width:100% !important;}"
            # メッシュのツールチップが下の観測点マーカーのクリックを奪わない
            # ようにする（他の地図と同じ扱い）
            ".leaflet-tooltip{pointer-events:none;width:max-content;"
            "max-width:240px;font-size:11.5px;line-height:1.5;padding:5px 8px;}"
            + POINT_LEGEND_CSS +
            "</style>"
        ))
        # 形（●＝一般国道／■＝高速自動車国道）の凡例は、この地図でも出す。
        # この地図の印は白く塗るので、凡例の印も白にする（plain=True）。
        fmap.get_root().html.add_child(folium.Element(
            point_legend_html(point_summary, plain=True)
        ))
        mesh_population.add_mesh_layer(
            fmap, geojson, metric, interactive=mesh_clickable
        )
        folium.Marker(
            [mainshock["epicenter_lat"], mainshock["epicenter_lon"]],
            icon=_epicenter_icon(), tooltip="本震の震源",
        ).add_to(fmap)
        add_map_size_fixer(fmap)
        # 選択で変わるもの（選択中のメッシュの枠と観測点の印）は、地図本体
        # ではなく feature_group_to_add に載せる。地図本体に入れると、選択の
        # たびに3,923件のメッシュを含む1.6MBのHTMLを作り直してブラウザへ
        # 送り直すことになり、クリックの反応が鈍くなる。
        overlay = folium.FeatureGroup(name="選択と観測点")
        # 選択中のメッシュは、同じ形・同じツールチップの図形を太枠で重ねる。
        # 枠だけのRectangleを乗せると、canvasの当たり判定は塗りの有無を見ない
        # ので、その枠が下のメッシュのクリックを奪って解除できなくなる。
        mesh_population.add_selection_layer(
            overlay,
            mesh_population.selected_geojson(
                summary, hour_label, day_label, selected,
                value_mode, residence_label,
            ),
            interactive=mesh_clickable,
        )
        build_plain_points_feature_group(
            point_summary, point_labels, selected_points,
            interactive=not mesh_clickable, into=overlay,
        )
        points_fg = overlay
        map_state = st_folium(
            fmap, height=MAP_HEIGHT_PX, width=550,
            feature_group_to_add=points_fg,
            returned_objects=[
                "last_object_clicked_tooltip", "last_object_clicked_count",
            ],
            key="mesh_pop_map_v1",
        )
        if st.button("選択をクリア", key="clear_mesh",
                     disabled=not (selected or selected_points)):
            st.session_state["selected_meshes"] = []
            st.session_state["selected_points_mesh"] = []
            st.session_state["_sel_version_mesh"] = sel_version + 1
            rerun_fragment()
        st.caption(
            (f"色は**{metric}**での、発災前後の平均人口の比です"
             if mesh_population.VALUE_MODES[value_mode] == "ratio"
             else f"色は**{metric}**の平均人口（{value_mode[3:-1]}）そのものです。"
                  "増減率に戻すと、どこで増えたか・減ったかが読めます。")
            + ((f"（発災前の平日{pre_days.get('平日', '—')}日 vs 発災後の平日"
                f"{post_days.get('平日', '—')}日）。" if day_label == "平日"
                else f"（発災前の休日{pre_days.get('休日', '—')}日 vs 発災後の"
                     f"休日{post_days.get('休日', '—')}日。日数が少ないので、"
                     "1日の事情が比にそのまま出ます）。")
               if mesh_population.VALUE_MODES[value_mode] == "ratio" else "")
            + f"いまクリックで選べるのは**{click_target}です**"
            "（メッシュ・観測点ともに1つずつ）。"
            "選んだものを右の図に重ねて出します（反映に1〜2秒かかります）。"
            "**選んでいるものは白い縁取りの上に図と同じ色の枠**で示します"
            "（メッシュは青緑、観測点は朱色）。"
            "白い縁取りは、塗りの色に関係なく輪郭が見えるようにするためです。"
            "観測点は白丸の印で、メッシュの色を隠さないよう"
            "異常度による描き分けはしていません"
            "（異常度で描き分けた地図は交通量のタブにあります）。"
            f"地図に出しているのは、全{meta['hours']}時点で配信された"
            f"**{len(geojson['features']):,}メッシュ**です"
            "（10人未満は配信されないため、時間帯によって出たり消えたりする"
            f"メッシュは同じ土台で比べられず、全{meta['meshes']:,}メッシュを"
            "描くと地図が重くなります）。"
        )

    click_count = map_state.get("last_object_clicked_count") if map_state else None
    if click_count is not None and click_count != st.session_state.get(
        "_last_click_count_mesh"
    ):
        st.session_state["_last_click_count_mesh"] = click_count
        tooltip = map_state.get("last_object_clicked_tooltip")
        # 当たり判定でも外してあるが、こちら側でも選んでいる種類だけ見る
        mesh = (mesh_population.mesh_from_tooltip(tooltip)
                if mesh_clickable else None)
        pid = (None if mesh_clickable
               else point_id_from_tooltip(tooltip, point_labels))
        if mesh is not None:
            if mesh in selected:
                selected.remove(mesh)
            else:
                if len(selected) >= MAX_SELECTED_MESHES:
                    selected.pop(0)
                selected.append(mesh)
            st.session_state["selected_meshes"] = selected
            st.session_state["_sel_version_mesh"] = sel_version + 1
            rerun_fragment()
        elif pid is not None:
            if pid in selected_points:
                selected_points.remove(pid)
            else:
                while len(selected_points) >= MAX_SELECTED_POINTS_ON_MESH_MAP:
                    selected_points.pop(0)
                selected_points.append(pid)
            st.session_state["selected_points_mesh"] = selected_points
            st.session_state["_sel_version_mesh"] = sel_version + 1
            rerun_fragment()

    with col_ts:
        render_mesh_timeseries(
            selected, selected_points, summary, meta, point_labels,
            quake_at, holidays, other_event_times,
        )


@tab_fragment
def render_mesh_compare_tab(mainshock: dict) -> None:
    """
    同じ集計を条件違いで2枚並べて見比べるタブ。

    「発災前と発災後」「夜間と昼間」「平日と休日」のように、どの2つを比べたい
    かは見る人によって変わる。左右それぞれに時間帯・日区分・出す値を持たせて、
    組み合わせを自由に選べるようにしている。

    2枚は folium の DualMap で、拡大・移動は連動する。見比べる図で縮尺が
    ずれると意味がないため。クリックでの選択はこのタブには置かない
    （どちらの地図を押したのかが返り値から分からず、取り違えるため）。
    """
    try:
        _mesh_compare_body(mainshock)
    except mesh_population.MeshDataUnavailable as e:
        st.error(f"モバイル空間統計の集計データを読めませんでした。\n\n{e}")
    except Exception as e:
        st.error(
            f"人口の比較のタブを描けませんでした（{type(e).__name__}: {e}）。"
            "ほかのタブは通常どおり動きます。"
            + _stale_module_hint(e)
        )


# 左右の既定の組み合わせ。左に発災前の人口、右に発災前後の増減率を出す
# （どこに人がいた場所で、どれだけ動いたかが1画面で読める）。
# 時間帯・日区分・居住地は左右で揃えておく。
MESH_COMPARE_DEFAULTS = (
    ("左", "人口（発災前）"),
    ("右", "増減率"),
)


def _mesh_compare_body(mainshock: dict) -> None:
    if not mesh_population.available():
        st.info(
            "モバイル空間統計の集計データが見つかりません。"
            "`python scripts/build_mesh_population.py` で作るか、"
            "非公開の置き場を `st.secrets` に設定してください。"
        )
        return

    meta = mesh_population.load_meta()
    summary = mesh_population.summary()
    st.caption(
        f"{meta['source']}（{meta['area']}・{meta['unit']}）。"
        f"収録は **{pd.Timestamp(meta['start']):%Y-%m-%d %H:%M} 〜 "
        f"{pd.Timestamp(meta['end']):%Y-%m-%d %H:%M}**。"
        "2枚の地図は拡大・移動が連動します。"
        "条件を選ぶと同じ場所どうしで見比べられます"
        "（クリックでの選択と時系列変化は「人口推計値×交通量」タブに"
        "あります）。"
    )

    hours = list(mesh_population.HOUR_METRICS)
    days = list(mesh_population.DAY_METRICS)
    modes = list(mesh_population.VALUE_MODES)
    residences = list(mesh_population.RESIDENCE_METRICS)

    columns = st.columns(2, gap="large")
    sides = []
    for (side, default_mode), col in zip(MESH_COMPARE_DEFAULTS, columns):
        with col:
            c_mode, c_hour = st.columns(2)
            c_day, c_res = st.columns(2)
            with c_mode:
                mode = st.selectbox(
                    "凡例を表示する値", modes,
                    index=modes.index(default_mode),
                    key=f"cmp_mode_{side}",
                )
            with c_hour:
                hour = st.selectbox(
                    "時間帯区分（夜間・昼間・全日）", hours,
                    index=hours.index(mesh_population.DEFAULT_HOUR),
                    key=f"cmp_hour_{side}",
                )
            with c_day:
                day = st.selectbox(
                    "平日・休日区分", days,
                    index=days.index(mesh_population.DEFAULT_DAY),
                    key=f"cmp_day_{side}",
                )
            with c_res:
                residence = st.selectbox(
                    "居住者・来訪者区分", residences,
                    index=residences.index(mesh_population.DEFAULT_RESIDENCE),
                    key=f"cmp_res_{side}",
                )
            label = mesh_population.metric_label(hour, day, residence)
            st.markdown(
                mesh_population.legend_html(label, mode), unsafe_allow_html=True
            )
        sides.append((hour, day, mode, residence, label))

    dual = folium_plugins.DualMap(
        location=[mainshock["epicenter_lat"], mainshock["epicenter_lon"]],
        zoom_start=MAP_ZOOM + 1, tiles=None, prefer_canvas=True,
        layout="horizontal",
    )
    n_meshes = 0
    for fmap, (hour, day, mode, residence, label) in zip(
        (dual.m1, dual.m2), sides
    ):
        folium.TileLayer(
            "OpenStreetMap", name="OpenStreetMap", opacity=0.45, control=False,
        ).add_to(fmap)
        geojson = mesh_population.mesh_geojson(
            summary, hour, day, MESH_MIN_HOURS, mode, residence
        )
        n_meshes = len(geojson["features"])
        mesh_population.add_mesh_layer(fmap, geojson, label)
        folium.Marker(
            [mainshock["epicenter_lat"], mainshock["epicenter_lon"]],
            icon=_epicenter_icon(), tooltip="本震の震源",
        ).add_to(fmap)
        add_map_size_fixer(fmap)
    # 触れているメッシュを両側で囲む。2枚目に付けるのは、その時点で
    # 1枚目のメッシュも作られていて、両方を辿れるため。
    MeshHoverLink(dual.m1).add_to(dual.m2)
    dual.get_root().header.add_child(folium.Element(
        "<style>.leaflet-tooltip{width:max-content;max-width:240px;"
        "font-size:11.5px;line-height:1.5;padding:5px 8px;}</style>"
    ))
    st_folium(
        dual, height=MESH_COMPARE_MAP_HEIGHT_PX, use_container_width=True,
        returned_objects=[], key="mesh_compare_map_v1",
    )
    st.caption(
        "左右で違うのは選んだ条件だけで、描いているメッシュ"
        f"（全{meta['hours']}時点で配信された{n_meshes:,}メッシュ）は"
        "同じです。増減率の区切りは15%・30%、人口の区切りは100・250・500・"
        "1,000人で、左右のどちらでも同じ色は同じ値を指します。"
    )


def render_mesh_timeseries(selected, selected_points, summary, meta,
                           point_labels: dict, quake_at, holidays,
                           other_event_times=()) -> None:
    """
    選択メッシュの人口推計値と、選択観測点の交通量を1つの図に重ねる。

    単位が違うので縦軸を分ける（左＝人、右＝台/時）。交通量は人口に合わせて
    1時間値の上下計で、平常時の線はメッシュ側だけに出す（8本になると
    どの線を追っているのか分からなくなるため。交通量の平常時との比較は
    「道路規制×交通量」タブが本体）。
    """
    if not selected and not selected_points:
        st.info(
            "地図のメッシュや観測点（どちらも1つ）をクリックすると、"
            "人口推計値と交通量がここに重ねて出ます。"
        )
        return

    fig = go.Figure()
    for mesh, color, bar_color in zip(
        selected, MESH_CHART_COLORS, MESH_BAR_COLORS
    ):
        frame = mesh_population.with_baseline(
            mesh_population.series_for(mesh, meta), pd.Timestamp(quake_at), holidays
        )
        # 人口は棒で出す。1時間ごとに「その時刻にそこに居た人数」を数えた値
        # なので、時点の量として読むほうが分かりやすい（交通量は流れなので線）。
        # 居住者・来訪者・内訳なしの3段に積む。積み上げの合計は総数になる。
        # 系列名は短くする。どのメッシュ・どの観測点かは図の見出しに出す
        # （名前に入れると凡例が長くなり、変なところで折り返す）。
        segments = (
            ("pop_rs", "居住者", bar_color[0]),
            ("pop_vi", "来訪者", bar_color[1]),
            ("pop_other", "内訳なし", bar_color[2]),
        )
        for rank, (column, seg_label, seg_color) in enumerate(segments, start=1):
            fig.add_trace(go.Bar(
                x=frame["datetime"], y=frame[column],
                marker=dict(color=seg_color, line=dict(width=0)),
                name=seg_label, legendrank=rank,
            ))
        # 平常時は、棒と同じ1時間ごとの水準なので階段（hv）で描く。
        # hvは「その点から次の点まで水平」に引くので、時刻そのままだと
        # 棒（時刻を中心に前後30分）から右へ30分ずれる。30分手前から引き、
        # 最後の棒の右端まで1点足して、段と棒の幅をそろえる。
        half = pd.Timedelta(minutes=30)
        step_x = list(frame["datetime"] - half) + [
            frame["datetime"].iloc[-1] + half
        ]
        step_y = list(frame["baseline"]) + [frame["baseline"].iloc[-1]]
        fig.add_trace(go.Scatter(
            x=step_x, y=step_y,
            mode="lines", line=dict(color=color, width=1.6, shape="hv"),
            name="人口 平常時", legendrank=4,
        ))

    # 横軸は交通量の図と同じ左端（本震の前日）から、モバイル空間統計の
    # 収録の終わりまで。人口は07-21から、交通量はこれより後まであるが、
    # 重ねて読む図なので両方が揃っている範囲に合わせる。
    x_from, x_to = TIMESERIES_DISPLAY_START, pd.Timestamp(meta["end"])
    if selected_points:
        hourly = load_observations("observations_hourly.parquet")
        for pid, color in zip(selected_points, [POINT_CHART_COLOR]):
            sub = hourly[
                (hourly["point_id"] == pid)
                & hourly["datetime"].between(x_from, x_to)
            ].sort_values("datetime")
            if sub.empty:
                continue
            # 上下を足して1本にする。平常時も同じ足し方（交通量のタブと
            # 同じ、同じ日区分・同じ時刻の平均）。
            fig.add_trace(go.Scatter(
                x=sub["datetime"],
                y=sub["baseline_mean_up"] + sub["baseline_mean_down"],
                yaxis="y2", mode="lines",
                line=dict(color=color, dash="dot", width=1),
                opacity=0.6, name="交通量 平常時", legendrank=6,
            ))
            fig.add_trace(go.Scatter(
                x=sub["datetime"],
                y=sub["traffic_up"] + sub["traffic_down"],
                yaxis="y2", mode="lines", line=dict(color=color, width=1.2),
                name="交通量", legendrank=5,
            ))

    for t in other_event_times:
        fig.add_vline(x=t, line_dash="dot", line_color="lightgray",
                      line_width=1, opacity=0.7)
    fig.add_vline(x=quake_at, line_dash="dot", line_color="black", line_width=2)
    fig.add_annotation(
        x=quake_at, y=1, yref="paper", yanchor="bottom",
        text="地震発生 16:27", showarrow=False, font=dict(size=11, color="black"),
    )
    fig.update_layout(
        # 交通量は夜間にほぼ0まで落ちる。人口の棒と同じ高さで伸ばすと、
        # 棒の側が読みにくくなるので600→500に戻す。
        height=500,
        # 居住地の区分は積む。棒どうしを少し空けて、1時間ごとの刻みが
        # 見えるようにする。
        barmode="stack", bargap=0.25,
        margin=dict(l=10, r=10, t=26, b=54),
        # 系列は6本までなので、凡例は1行に並ぶ。項目の幅を揃えて、
        # 変なところで折り返さないようにする。
        legend=dict(orientation="h", yref="container", yanchor="bottom", y=0.012,
                    xanchor="left", x=0, font=dict(size=10),
                    # 追加した順に並べる。既定（reversed）だと交通量が先に
                    # 来て、人口の3段が後ろに回り、読む順と食い違う。
                    traceorder="normal", itemwidth=30),
        yaxis=dict(title="人口推計値（人）", rangemode="tozero"),
        yaxis2=dict(
            title="交通量（台/時・上下計）", overlaying="y", side="right",
            rangemode="tozero", showgrid=False,
        ),
        xaxis=dict(title="", range=[x_from, x_to]),
        hovermode="x unified",
        # 系列名はplotlyの既定だと15文字で切られ、メッシュ番号＋市町村名が
        # 途中で消える。名前を切らずに出す（吹き出しはその分広がる）。
        hoverlabel=dict(namelength=-1, font=dict(size=11)),
    )
    st.plotly_chart(fig, use_container_width=True, key="mesh_pop_ts")
    st.caption(
        "**左の縦軸が人口推計値（人・棒）、右の縦軸が交通量（台/時・上下計・線）です**。"
        "人口はその時刻にそこに居た人数なので棒、交通量は流れなので線にしています。"
        "交通量は人口に合わせて1時間値を使っています。"
        f"表示期間は交通量の図と同じ左端（本震の前日 {x_from:%m-%d %H:%M}）から、"
        f"モバイル空間統計の収録の終わり（{x_to:%m-%d %H:%M}）まで"
        "（人口はこれより前、交通量はこれより後までデータがありますが、"
        "重ねて読む図なので両方が揃っている範囲に合わせています）。"
        "**棒は居住地で塗り分けています**（濃い色＝そのメッシュのある市区町村の"
        "居住者、薄い色＝来訪者、いちばん薄い色＝居住地別が10人未満で配信され"
        "なかった分）。3つを積むと総数になります。"
        "**平常時は、人口が棒と同じ刻みの階段線、交通量が点線**です。"
        "交通量の平常時は交通量のタブと同じもの（地震前の同じ日区分・同じ時刻の"
        "平均。曜日ごとに各8日分）で、上下を足しています。"
        "人口の平常時＝発災前（"
        f"{pd.Timestamp(meta['start']):%m-%d} 〜 本震まで）の、同じ日区分"
        "（平日／土／日祝）・同じ時刻の平均。**発災前が8日分しかないため、"
        "人口側は曜日ごとではなく3区分で平均し、土・日祝はそれぞれ1日分の値です**"
        "（交通量側とは平常時の作り方が違います）。"
        "線が切れているのは、その時間帯に10人未満で配信されなかったところです。"
    )


def main():
    st.markdown(
        "<style>"
        'iframe[title="streamlit_folium.st_folium"] { width: 100% !important; }'
        # 本題（地図・時系列・異常検知）が早く視界に入るよう、上部の余白を詰める。
        # ただしStreamlit標準のヘッダー（Deploy/共有/メニュー）は position:fixed の
        # 高さ60pxで、本文の余白がそれより小さいとタイトルがヘッダーに潜って欠ける。
        # ヘッダー自体を低くするとツールバーのボタン（下端が固定で52px）がはみ出すため、
        # ヘッダーには触れず、その高さを必ず上回る余白にする（既定の6remよりは詰める）。
        ".stMainBlockContainer, .block-container { padding-top: 4.2rem !important; }"
        # 見出しは既定だと大きすぎて、それだけで1画面を使ってしまう
        ".dash-title { font-size:1.55rem; font-weight:700; line-height:1.3; margin:0 0 2px 0; }"
        ".dash-lead  { font-size:0.92rem; line-height:1.5; margin:0 0 6px 0; }"
        ".dash-src   { font-size:0.78rem; color:#5b6570; line-height:1.5; margin:0; }"
        # 粒度ラジオは4択。横並びのままだと 3+1 で折り返して座りが悪いので、
        # 2列のグリッドに固定して 2x2 にする。
        #   行 = 合計 / 車種別、列 = 5分間値 / 1時間値 の対応で読める。
        # 画面が狭いときは列が縦に積まれるので、グリッドは広い画面だけに掛ける。
        "@media (min-width: 901px){"
        '[data-testid="stRadio"] [role="radiogroup"]'
        "{display:grid;grid-template-columns:1fr 1fr;gap:0 8px;}}"
        # 観測点の選択と粒度の操作行を画面上部に貼り付ける。地図で観測点を
        # 選び直すたびにスクロールで戻る必要があったため。
        # sticky はウィジェット本体ではなく、それを包む element-container ではなく
        # さらに外側の「行」に掛ける必要がある。ウィジェット直下の
        # element-container は中身と同じ高さで、sticky が動く余地が無いため
        # （実測: どちらも 737px で一致し、貼り付いても位置が変わらない）。
        # スクロールしているのは section.main（overflow:auto）なので、
        # その中で sticky が効く。標準ヘッダー（position:fixed・高さ60px）に
        # 隠れないよう top を下げる。
        '[data-testid="stHorizontalBlock"]:has([data-testid="stMultiSelect"])'
        "{position:sticky;top:3.9rem;z-index:20;"
        "background-color:var(--background-color,#ffffff);"
        "padding:6px 2px 4px 2px;border-bottom:1px solid rgba(130,130,130,0.25);}"
        "@media (prefers-color-scheme: dark){"
        '[data-testid="stHorizontalBlock"]:has([data-testid="stMultiSelect"])'
        "{background-color:var(--background-color,#0e1117);}}"
        # 画面が狭いと列が縦に積まれて操作行が高くなり、貼り付けると
        # 画面の大半を占めてしまうので、その場合は貼り付けない。
        "@media (max-width: 900px){"
        '[data-testid="stHorizontalBlock"]:has([data-testid="stMultiSelect"])'
        "{position:static;border-bottom:none;}}"
        # ページ切り替えのタブ。既定は文字の下に細い線が付くだけで、
        # どれが選べるのか・いまどれを見ているのかが分かりにくい。
        # 選んでいないタブを一段引っ込め、選んでいるタブを手前に浮かせる。
        # 入れ子のタブ（列定義書）にも同じ見た目が掛かるが、あちらも
        # 切り替えの操作なので揃っていて困らない。
        #
        # タブのDOMはStreamlitの版で変わる。手元の1.37は baseweb の属性
        # （tab-list の中のbutton）、公開先の新しい版は role="tablist" と
        # data-testid="stTab" になる。片方だけを指定すると、もう片方では
        # 何も掛からず既定の見た目に戻ってしまうので、両方に当てる。
        # 新しい版はタブ自身に padding:0 とクラスが当たっているため、
        # competing する指定には !important を付ける。
        'div[data-baseweb="tab-list"],[data-testid="stTabs"] [role="tablist"]'
        "{gap:6px;border-bottom:1px solid rgba(130,130,130,0.35);"
        "padding:4px 2px 0 2px;}"
        # 新しい版の下線は tablist の ::after（2px）で描かれている
        '[data-testid="stTabs"] [role="tablist"]::after{display:none;}'
        'div[data-baseweb="tab-list"]>button,[data-testid="stTab"]'
        "{border:1px solid rgba(130,130,130,0.35) !important;"
        "border-bottom:none !important;"
        "border-radius:7px 7px 0 0 !important;padding:6px 14px !important;"
        "margin-bottom:-1px;background:rgba(130,130,130,0.10) !important;"
        "box-shadow:inset 0 -3px 5px -4px rgba(0,0,0,0.45);"
        "transition:background .12s, transform .12s;}"
        'div[data-baseweb="tab-list"]>button:hover,[data-testid="stTab"]:hover'
        "{background:rgba(130,130,130,0.18) !important;}"
        # 選択中はページと地続きに見せる（下線を消して手前に出す）
        'div[data-baseweb="tab-list"]>button[aria-selected="true"],'
        '[data-testid="stTab"][aria-selected="true"]'
        "{background:var(--background-color,#ffffff) !important;"
        "border-color:rgba(130,130,130,0.55) !important;font-weight:700;"
        "box-shadow:0 -2px 6px -2px rgba(0,0,0,0.30);transform:translateY(-1px);}"
        "@media (prefers-color-scheme: dark){"
        'div[data-baseweb="tab-list"]>button[aria-selected="true"],'
        '[data-testid="stTab"][aria-selected="true"]'
        "{background:var(--background-color,#0e1117) !important;}}"
        # 旧い版の下線（赤いハイライト）は枠と二重になるので消す
        'div[data-baseweb="tab-highlight"],div[data-baseweb="tab-border"]'
        "{display:none;}"
        "</style>",
        unsafe_allow_html=True,
    )
    data_missing = not os.path.exists(os.path.join(DATA_DIR, "observations.parquet"))
    if data_missing:
        st.error(
            "data/observations.parquet が見つかりません。先に `python fetch_and_prepare.py` を実行してください。"
        )
        st.stop()

    # 異常検知（zスコア・地図の色分け）は1時間値ベースで定義する。
    # 時系列図の実績は既定で5分間値を使う（load_observationsで別途読み込む）。
    observations = load_observations("observations_hourly.parquet")
    quake_info = load_quake_info()
    regulations = load_regulations()
    mlit = load_mlit_regulations()
    nexco = load_nexco_regulations()
    mainshock = quake_info["mainshock"]
    quake_at = pd.Timestamp(mainshock["occurred_at"]).tz_localize(None)
    other_event_times = [
        pd.Timestamp(e["occurred_at"]).tz_localize(None)
        for e in quake_info.get("events", [])
        if e.get("eid") != mainshock.get("eid")
    ]

    # 異常検知（地図の色分けの元）は発災後1週間だけを対象にする。
    # 平常時は地震前の同じ日区分から作っているので、お盆のように
    # そもそも交通の出方が違う時期を混ぜると「平常時との差」が
    # 災害の影響なのか行事の影響なのか区別できなくなる。
    anomaly_end = quake_at + ANOMALY_WINDOW
    post = observations[
        observations["is_post_quake"] & (observations["datetime"] < anomaly_end)
    ]
    point_summary = build_point_summary(post, observations)

    if "selected_points" not in st.session_state:
        st.session_state["selected_points"] = _default_selection(point_summary)

    INTENSITY_LABELS = {1: "1", 2: "2", 3: "3", 4: "4", 5: "5弱", 6: "5強", 7: "6弱", 8: "6強", 9: "7"}
    n_events = len(quake_info.get("events", []))
    min_intensity_label = INTENSITY_LABELS.get(quake_info.get("events_min_intensity"), "?")
    period_start = quake_info.get("events_period_start", "?")
    period_end = quake_info.get("events_period_end", "?")

    # 以前はマグニチュード・最大震度・観測点数・取得件数・異常件数を
    # st.metric 5枚で並べていたが、縦を取るわりに読む頻度が低い。
    # 地震の諸元はタイトルに、データの規模はリード文に畳み込み、
    # 観測点数は地図の見出しへ移した（異常件数は一覧タブにあるので出さない）。
    raw_5min = load_traffic_archive("traffic_raw.parquet")
    n_points = len(point_summary)
    scale_note = ""
    if not raw_5min.empty:
        scale_note = (
            f"（対象範囲の{n_points}点で全{len(raw_5min):,}件、"
            f"{raw_5min['datetime'].max():%Y-%m-%d %H:%M}時点）"
        )
    st.markdown(
        '<div class="dash-title">熊本地震（2026-07-28 M{m}・最大震度{i}）'
        '交通量変化ダッシュボード</div>'
        '<div class="dash-lead">'
        '<b><a href="https://www.jartic-open-traffic.org/" target="_blank">'
        'JARTIC 交通量オープンデータ</a></b> の常設トラカン交通量（5分間値・1時間値）'
        '{note}と、<b>モバイル空間統計®</b> の推計人口（500mメッシュ・1時間）を'
        '主データに、平常時と比べて交通量・人口がどれだけ外れたかを'
        '可視化しています。'
        '通行規制と地震情報を重ね合わせ、変化の背景を追えるようにしています。'
        "</div>".format(
            m=mainshock["magnitude"],
            i=mainshock["max_intensity"] or "?",
            note=scale_note,
        ),
        unsafe_allow_html=True,
    )

    archive_period = ""
    if not raw_5min.empty and "datetime" in raw_5min.columns:
        archive_period = (
            f"交通量アーカイブ: {raw_5min['datetime'].min():%Y-%m-%d %H:%M}"
            f" 〜 {raw_5min['datetime'].max():%Y-%m-%d %H:%M}　｜　"
        )
    st.markdown(
        '<div class="dash-src">'
        f"{archive_period}"
        f"異常検知の対象期間: 本震（{period_start}）〜 {period_end}"
        f"（本震+3週間または現在時刻の早い方）　｜　"
        f"データ生成: {quake_info.get('generated_at', '不明')}<br>"
        # 並びは 交通量 → 人口 → 通行規制（道路種別の階層順：全体をまとめた
        # 配布データ、高速道路、直轄国道、県が管理する補助国道・県道・
        # 市町村道）→ 地震情報。県の公開JSONには補助国道（218/219/265/
        # 266/324/442/443/445号）も入るので、そう書く。
        "データ源: "
        '<a href="https://www.jartic-open-traffic.org/" target="_blank">JARTIC 交通量オープンデータ</a>'
        ' ／ <a href="https://mobaku.jp/" target="_blank">'
        'モバイル空間統計®（推計人口）</a>'
        ' ／ <a href="https://www.mlit.go.jp/road/saigai/r8kumamoto/index.html"'
        ' target="_blank">国土交通省「通れる道マップ」（通行規制情報）</a>'
        ' ／ <a href="https://www.w-nexco.co.jp/" target="_blank">NEXCO西日本（高速道路の規制）</a>'
        ' ／ <a href="https://www.qsr.mlit.go.jp/kumamoto/" target="_blank">熊本河川国道事務所（直轄国道の規制）</a>'
        ' ／ <a href="https://portal.bousai.pref.kumamoto.jp/?p=traffic" target="_blank">防災情報くまもと（補助国道・県道・市町村道の規制）</a>'
        ' ／ <a href="https://www.jma.go.jp/jma/menu/20260728_kumamoto_jishin.html" target="_blank">気象庁（地震情報）</a>'
        # 交通量API利用規約 第5条1項が求める出典。「エンドユーザーが当該
        # サービスを利用する度に確認できる位置」に出す必要があるので、
        # タブより前・タイトル直下の常時表示にしている。
        # 第5条2項（加工して利用する場合）と第8条2項（利用者が一切の責任を
        # 負う旨の明記）もここでまとめて満たす。
        f'<br>{JARTIC_TERMS_NOTICE}'
        "</div>",
        unsafe_allow_html=True,
    )

    MAXI_DISPLAY = {
        "1": "1", "2": "2", "3": "3", "4": "4",
        "5-": "5弱", "5+": "5強", "6-": "6弱", "6+": "6強", "7": "7",
    }
    # 地震一覧は参照用なので折りたたむ（開いたままだと本題が画面の下に押し出される）
    # 見出しに本震だけの諸元（震源の深さ・震源地）を併記すると、
    # 複数地震の一覧であることと食い違うので入れない。
    with st.expander(
        f"最大震度{min_intensity_label}以上を観測した地震の一覧（{n_events}件）"
    ):
        events_rows = []
        for e in quake_info.get("events", []):
            dt = datetime.fromisoformat(e["occurred_at"])
            events_rows.append({
                "発生時刻": dt.strftime("%Y年%m月%d日%H時%M分"),
                "震央地名": e["epicenter_name"],
                "マグニチュード": e["magnitude"],
                "最大震度": MAXI_DISPLAY.get(e["max_intensity"], e["max_intensity"] or "-"),
                "震度": f"https://www.jma.go.jp/bosai/map.html#&contents=estimated_intensity_map&id={dt.strftime('%Y%m%d%H%M')}",
            })
        st.dataframe(
            pd.DataFrame(events_rows),
            hide_index=True,
            use_container_width=True,
            column_config={
                "震度": st.column_config.LinkColumn("震度", display_text="推計震度分布図"),
            },
        )
        st.caption(
            "出典: [気象庁](https://www.jma.go.jp/jma/menu/20260728_kumamoto_jishin.html)。"
            "推計震度分布図は地震発生直後に発表されたもの（発表がない地震ではリンク先に情報がない場合があります）。"
        )

    # 通れる道マップ版を先に置く。規制の収録が国の配信で揃っており、
    # 高速道路の緊急車両の通行可能区間まで区間ごとの線で入るため、
    # こちらを主に見てもらう。県のJSON＋PDF転記の版は(旧版)に下げた。
    tab_mlit, tab_mesh, tab_mesh_cmp, tab_overview, tab_dl = st.tabs(
        [
            "道路規制×交通量",
            "人口推計値×交通量",
            "人口推計値の時点比較",
            "(旧版)道路規制×交通量",
            "データダウンロード",
        ]
    )

    # ------------------------------------------------------------------
    # (旧版)道路規制×交通量タブ
    # 県の公開JSON＋PDFからの転記で集めた規制の版
    # ------------------------------------------------------------------
    with tab_overview:
        if post.empty:
            st.info("地震発生後のデータがまだありません。")
        else:
            point_labels = build_point_labels(point_summary)
            coords_by_id = point_summary.set_index("point_id")[
                ["point_lat", "point_lon"]
            ].to_dict("index")
            selected_points = st.session_state["selected_points"]
            sel_version = st.session_state.get("_sel_version", 0)

            def _format_point(pid: str) -> str:
                c = coords_by_id[pid]
                return f"{point_labels[pid]}（{c['point_lat']:.3f}N, {c['point_lon']:.3f}E）"

            # 観測点の選択と粒度の切り替えを1行にまとめる。観測点を比べるときに
            # 何度も触るのはこの2つなので、同じ行に置いてまとめて画面に貼り付ける
            # （下のCSSで sticky）。粒度ラジオが地図より前に作られる点も維持する。
            col_select, col_view, col_range, col_clear = st.columns([3, 2, 1.3, 1])
            with col_select:
                picked = st.multiselect(
                    f"観測点の選び方：このプルダウンから選ぶか、地図上の丸いマーカーをクリック"
                    f"（最大{MAX_SELECTED_POINTS}地点まで並べて比較できます）",
                    options=list(point_labels.keys()),
                    default=[p for p in selected_points if p in point_labels],
                    max_selections=MAX_SELECTED_POINTS,
                    format_func=_format_point,
                    key=f"point_select_{sel_version}",
                )
            with col_view:
                view_names = list(TIMESERIES_VIEWS.keys())
                saved_view = st.session_state.get("_timeseries_view_choice", view_names[0])
                view = st.radio(
                    "交通量の集計方法",
                    view_names,
                    index=view_names.index(saved_view) if saved_view in view_names else 0,
                    horizontal=True,
                    key="timeseries_view",
                )
                st.session_state["_timeseries_view_choice"] = view
                cfg = TIMESERIES_VIEWS[view]
            with col_range:
                # 粒度ラジオと同じ理由で地図より前に作る（地図クリックの
                # st.rerun() で中断されるとウィジェットの状態が落ちるため）。
                range_names = list(TIMESERIES_RANGES.keys())
                saved_range = st.session_state.get(
                    "_timeseries_range_choice", TIMESERIES_DEFAULT_RANGE
                )
                range_name = st.selectbox(
                    "交通量のグラフの表示期間",
                    range_names,
                    index=(
                        range_names.index(saved_range) if saved_range in range_names
                        else range_names.index(TIMESERIES_DEFAULT_RANGE)
                    ),
                    key="timeseries_range",
                )
                st.session_state["_timeseries_range_choice"] = range_name
            with col_clear:
                st.write("")
                if st.button("選択をクリア", disabled=not selected_points):
                    st.session_state["selected_points"] = []
                    st.session_state["_sel_version"] = sel_version + 1
                    st.rerun()

            # プルダウンの選択は、この実行のなかでそのまま採用する。
            # 以前は session_state に入れて st.rerun() していたが、選択1回あたり
            # スクリプトが2回走り（実測: 96ms差で連続実行）、体感で倍の待ち時間に
            # なっていた。地図・時系列はこの下で描かれるので、picked を
            # そのまま使えば1回の実行で反映できる。
            if picked != selected_points:
                selected_points = picked
            st.session_state["selected_points"] = picked

            # 地図と時系列が近すぎて見分けづらいので、列の間隔を広く取る
            col_map, col_ts = st.columns([2, 3], gap="large")

            # 粒度のラジオは地図より先に作る。地図クリックの処理は st.rerun() で
            # スクリプトを中断するため、そこより後ろでウィジェットを作ると
            # その実行では作られず、Streamlit側の状態が落ちて既定値（5分間値）に
            # 戻ることがある（画面のラジオは1時間値のままなのに図だけ5分値になる）。
            # あわせて、選択内容をウィジェット以外のキーにも控えておき、
            # 状態が落ちても index で復元できるようにする。
            with col_ts:
                st.subheader("選択観測点の交通量の時系列変化（平常時 vs 観測実績）")

            with col_map:
                st.subheader("常時観測点別の異常度 × 通行規制")
                _now = _now_jst()
                # 色（規制の区分）と線の形（規制中か解除済みか）は独立した軸で、
                # 破線は青灰・赤・橙のどれにも、県feedにも直轄国道にも同じように
                # 掛かる。以前は「破線は解除済み」を地震前の規制の行に置いていて、
                # その区分だけの決まりごとに見えていたので、行を分けた。
                n_post = sum(
                    1 for r in regulations if _regulation_is_post_quake(r, quake_at)
                )
                n_pre = len(regulations) - n_post
                n_ended = sum(1 for r in regulations if _regulation_is_ended(r, _now))
                # 凡例は3行に圧縮する。見出しを独立行にすると1項目ごとに
                # 行が増えて余白ばかりになるので、見出しも同じ行に入れて
                # 「分類: 項目 項目 …」の形にし、文字も小さくする。
                # 件数はここには入れない。全ての項目に付けられるわけでは
                # なく（色は区分ごとの数を数えていない）、付いたり付かなかったり
                # すると何の数なのかがかえって分からなくなる。件数は地図の下の
                # データソースの表と、リンク先の一覧のほうに置いている。
                n_points = len(point_summary)
                _mlit_items = (mlit or {}).get("items", [])
                n_mlit_line = sum(1 for i in _mlit_items if i.get("path"))
                n_mlit_point = sum(
                    1 for i in _mlit_items
                    if not i.get("path") and i.get("affected_point_codes")
                )
                # 線も▲も出せない規制。PDFが区間をIC名で示していない
                # （キロポストや地名だけの）ものがこれにあたる。
                n_mlit_spot = sum(
                    1 for i in _mlit_items if i.get("point") and not i.get("path")
                )
                n_ended += sum(
                    1 for i in _mlit_items if i.get("path") and i.get("end_timestamp")
                )

                def _sw(style: str) -> str:
                    """凡例の色見本。縦位置を文字に揃える。"""
                    return f'<span style="display:inline-block;vertical-align:middle;{style}"></span>'

                bar = "width:20px;height:5px;"
                color_items = [
                    f'{_sw(bar + "background:#e60000;")} 全面/車両通行止め',
                    f'{_sw("width:20px;height:4px;background:#e67e22;")} 片側交互など',
                ]
                if n_pre:
                    color_items.append(
                        f'{_sw("width:20px;height:3px;background:#5b7c99;opacity:0.55;")}'
                        ' 地震前からの規制（工事・過去の災害）'
                    )
                if n_mlit_line:
                    # 規制区間の端点となるICも点で置いている。
                    color_items.append(
                        f'{_sw("width:11px;height:11px;border-radius:50%;"
                              "background:#fff;border:2px solid #333;")}'
                        ' 規制区間の端点（IC）'
                    )
                # 記号ごとに意味を1つにする。色＝規制の区分、⊗＝地震後に
                # 始まったかどうか、実線/破線＝いまの状態、線か⊗単独か＝
                # 場所が区間で示されているか地点でしか分からないか。
                shape_items = [
                    '<span style="color:#555;font-weight:700;">⊗</span>'
                    ' 地震後に始まった規制（色は上の区分）',
                    f'{_sw("width:20px;height:4px;background:#95a5a6;")} 実線・外円が実線は規制中',
                ]
                if n_ended:
                    shape_items.append(
                        f'{_sw("width:20px;height:0;border-top:4px dashed #95a5a6;")}'
                        ' 破線・外円が点線は解除済み'
                    )
                if n_mlit_spot:
                    shape_items.append(
                        '区間は線、地点だけ分かるものは⊗のみ（位置は概略）'
                    )
                if n_mlit_point:
                    shape_items.append(
                        '<span style="color:#b00000;font-weight:900;">▲</span>'
                        ' 直轄国道で線形が作れないもの（該当観測点のすぐ上）'
                    )
                # 形（道路の種別）の凡例は地図の中に置いてあるのでここには出さない
                rows = [
                    ("規制の色", color_items),
                    ("規制の印と状態", shape_items),
                    ("観測点", [point_z_legend_html(point_summary)]),
                ]
                legend_html = "".join(
                    '<div style="display:flex;flex-wrap:wrap;gap:1px 12px;align-items:center;">'
                    f'<b style="white-space:nowrap;">{head}:</b>'
                    + "".join(f'<span style="white-space:nowrap;">{it}</span>' for it in items)
                    + "</div>"
                    for head, items in rows
                )
                base_map = build_base_map(
                    point_summary, mainshock, regulations, mlit, point_labels,
                    nexco=nexco,
                )
                add_map_size_fixer(base_map)
                points_fg = build_points_feature_group(
                    point_summary, point_labels, selected_points
                )
                map_state = st_folium(
                    # 縦を詰めて右側の2図と高さを近づける（750 -> 620 -> 496）
                    base_map, height=MAP_HEIGHT_PX, width=550,
                    feature_group_to_add=points_fg,
                    # 座標ではなく「何をクリックしたか」で判定する。
                    #   last_object_clicked_tooltip … クリックされた図形のツールチップ本文。
                    #        これで観測点マーカーかどうかを確実に見分けられる
                    #   last_object_clicked_count   … 図形クリックのたびに増える通し番号。
                    #        同じマーカーを続けて押しても値が変わるので反応する
                    returned_objects=[
                        "last_object_clicked_tooltip", "last_object_clicked_count",
                    ],
                    key="quake_map_v8",
                )
                # 凡例は地図の下に置く。上にあると地図そのものが下へ押し出され、
                # 開いた直後に地図の全体が見えなくなる。凡例は「地図を見て
                # 分からなかったときに目を落とす先」なので下で足りる。
                # （HTMLを出すだけでウィジェットではないので、地図より後に
                # 書いても st.rerun でstateが落ちる問題には掛からない）
                st.markdown(
                    '<div style="font-size:0.79rem;line-height:1.45;margin:4px 0 6px 0;">'
                    f"{legend_html}</div>",
                    unsafe_allow_html=True,
                )
                # 注記は、まずどこから何を集めたのかを表で示す。
                # 以前は経路①②③を文章で並べていたが、3経路に増えて
                # 読み通せない長さになったので、出典と収集済みデータへの
                # リンクを表に畳み、残りの補足だけを短く下に置く。
                st.markdown(
                    render_source_table(regulations, n_post, mlit, nexco),
                    unsafe_allow_html=True,
                )
                st.caption(
                    "①には県が管理する補助国道（国道219号など）も入ります。"
                    "①は始点・終点の座標だけなので[OSRM](https://project-osrm.org/)で"
                    "道路網にスナップ、②③のPDFには座標がないため区間はOSMのIC座標から"
                    "線形を復元しています（キロポストだけのものは地名からの概略位置）。"
                    "描き分けは3つに共通で、色が規制の区分、破線が解除済みです。"
                    "地図の右下の「≡ レイヤ」（マウスを載せると開きます）で、"
                    "道路の管理者×状態ごとに表示を"
                    "切り替えられます（既定は規制中のみ）。"
                    "②③で転記した規制の一覧（区間・期間・根拠にした報・"
                    f"線形の作り方）は[こちら]({REPO_URL}/blob/main/docs/regulations.md)にまとめています。"
                )
                _emergency = emergency_access_note(nexco)
                if _emergency:
                    st.caption(_emergency)
                st.caption(
                    "規制のアーカイブに記録が残っている最も古い時点は "
                    f"**{regulation_archive_start() or '本震の翌日'}**（本震より後）で、"
                    "それ以前に解除された規制は、県管理道路であっても残っていません。"
                    "①には2020年7月豪雨など今回の地震と無関係な長期規制も含まれるため、"
                    "本震（16:27）以降に始まったかどうかで分けています。"
                )
                st.caption(point_marker_caption(anomaly_end))

            # クリックされたのが観測点マーカーのときだけ選択を切り替える。
            # 通行規制の線などを押した場合はツールチップが一致しないので何も起きない
            # （ツールチップは表示されるので、押した本人の意図とも合う）。
            click_count = map_state.get("last_object_clicked_count") if map_state else None
            if click_count is not None and click_count != st.session_state.get("_last_click_count"):
                st.session_state["_last_click_count"] = click_count
                pid = point_id_from_tooltip(
                    map_state.get("last_object_clicked_tooltip"), point_labels
                )
                if pid is not None:
                    selected = list(st.session_state["selected_points"])
                    if pid in selected:
                        selected.remove(pid)
                    else:
                        if len(selected) >= MAX_SELECTED_POINTS:
                            selected.pop(0)
                        selected.append(pid)
                    st.session_state["selected_points"] = selected
                    # セレクタ側のウィジェット状態は古くなるので、キーを変えて作り直す
                    st.session_state["_sel_version"] = sel_version + 1
                    st.rerun()

            with col_ts:
                # 選択中の観測点に掛かっている直轄国道の通行止めを帯で示す
                mlit_bands = []
                for pid in selected_points:
                    row = point_summary[point_summary["point_id"] == pid]
                    if row.empty:
                        continue
                    for item in mlit_regulations_for_point(mlit, row["point_code"].iloc[0]):
                        start = _parse_reg_time(item.get("start_timestamp"))
                        end = _parse_reg_time(item.get("end_timestamp")) or observations["datetime"].max()
                        if start is None:
                            continue
                        mlit_bands.append({
                            "start": start, "end": end,
                            "label": (
                                f"{item['route_name']}（{item['section']}）{item['content']}"
                            ),
                        })
                ts_obs = load_observations(cfg["file"])
                last_at = (
                    ts_obs["datetime"].max() if not ts_obs.empty
                    else quake_at + pd.Timedelta(days=7)
                )
                render_timeseries(
                    ts_obs,
                    selected_points, quake_at, other_event_times, point_labels,
                    quake_info.get("hourly_baseline_windows"),
                    unit_label=cfg["unit_label"],
                    extra_note=cfg["note"],
                    mlit_bands=mlit_bands,
                    baseline_daytypes=quake_info.get("hourly_baseline_daytypes"),
                    series_mode=cfg.get("series", "total"),
                    x_range=TIMESERIES_RANGES[range_name](quake_at, last_at),
                )

    # ------------------------------------------------------------------
    # データダウンロードタブ
    # ------------------------------------------------------------------
    with tab_dl:
        st.subheader("アーカイブデータのダウンロード")
        st.caption(
            "JARTICの5分値は過去1ヶ月・1時間値は過去3ヶ月しか遡れず、通行規制は解除されると"
            "ポータルの一覧から消えてしまいます。このダッシュボードは取得した分を追記専用で"
            "蓄積しているため、ここから取得済みの全期間をCSVで取り出せます。"
        )
        # 配布するCSVは「本機能により取得した情報を用いた資料」にあたるので、
        # 交通量API利用規約 第5条2項の出典をダウンロード箇所にも出す。
        st.caption(f"<b>出典</b>: {JARTIC_TERMS_NOTICE}", unsafe_allow_html=True)

        downloads = [
            (
                "5分間交通量（生データ・アーカイブ全期間）",
                load_traffic_archive("traffic_raw.parquet"),
                "kumamoto_traffic_5min_archive.csv",
                "観測点（常時観測点コード）×日時ごとの上り・下り交通量。車種別の内訳列も含みます。",
            ),
            (
                "1時間交通量（生データ・アーカイブ全期間）",
                load_traffic_archive("traffic_hourly.parquet"),
                "kumamoto_traffic_hourly_archive.csv",
                "同じ観測点の1時間値。平常時（日区分ごとに8日分）の母集団もこのデータから作っています。",
            ),
        ]
        regs_df, hist_df = load_regulations_archive()
        downloads += [
            (
                "通行規制 一覧（アーカイブ全期間）",
                regs_df,
                "kumamoto_road_regulations_archive.csv",
                "1件1行。初回・最終確認日時と、まだポータルに載っているか（still_listed）を含みます。"
                "経路の座標列は列数が膨大になるためCSVには含めていません（`data/archive/regulations_archive.json` にあります）。",
            ),
            (
                "通行規制 状態変化の履歴",
                hist_df,
                "kumamoto_road_regulations_history.csv",
                "規制内容や終了日時が変わった時点ごとに1行。"
                "全面通行止め → 片側交互通行止め → 解除 といった推移を追えます。",
            ),
            (
                "異常検知の入力データ（1時間値＋平常時＋zスコア）",
                observations,
                "kumamoto_observations_hourly.csv",
                "地図の色分けの根拠になっている表そのものです。",
            ),
        ]

        # --- データセット一覧（表形式）---
        # 以前は列定義書を先に置いていたが、「上記すべてのCSV」と書いても
        # その時点でどんなCSVがあるか示されていなかった。先に一覧を出す。
        sheet_of = {fname: sheet for fname, sheet, _, _ in DATA_DICTIONARY}
        col_w = [2.4, 3.8, 1.9, 1.7]
        head = st.columns(col_w)
        for col, label in zip(head, ["データセット", "内容", "規模", "ダウンロード"]):
            col.markdown(f"<div style='font-weight:600;font-size:0.85rem;'>{label}</div>",
                         unsafe_allow_html=True)
        st.markdown(
            "<hr style='margin:2px 0 6px 0;border:none;border-top:1px solid #ccc;'>",
            unsafe_allow_html=True,
        )
        for title, df, fname, desc in downloads:
            c1, c2, c3, c4 = st.columns(col_w)
            with c1:
                st.markdown(f"**{title}**")
                st.caption(f"`{fname}`  \n列定義書のシート: **{sheet_of.get(fname, '—')}**")
            c2.caption(desc)
            if df is None or df.empty:
                c3.caption("まだデータがありません")
                c4.button("CSV", disabled=True, key=f"dl_{fname}_disabled")
            else:
                csv_bytes = to_csv_bytes(df)
                scale = f"{len(df):,} 行 / {len(csv_bytes) / 1024:,.0f} KB"
                if "datetime" in df.columns:
                    scale += (
                        f"  \n期間: {df['datetime'].min():%Y-%m-%d %H:%M}"
                        f" 〜 {df['datetime'].max():%Y-%m-%d %H:%M}"
                    )
                c3.caption(scale)
                c4.download_button(
                    "CSV", csv_bytes, file_name=fname, mime="text/csv",
                    key=f"dl_{fname}", use_container_width=True,
                )
            st.markdown(
                "<hr style='margin:4px 0;border:none;border-top:1px solid #eee;'>",
                unsafe_allow_html=True,
            )

        # --- 列定義書（データディクショナリ） ---
        actual_columns = {
            fname: list(df.columns)
            for _, df, fname, _ in downloads
            if df is not None and not df.empty
        }
        dict_df, dict_notes = build_data_dictionary(actual_columns)
        st.markdown("**CSVの列定義書（各列の意味・単位）**")
        st.caption(
            f"上の表の{len(DATA_DICTIONARY)}つのCSVについて、列の並び順・単位・意味をまとめたExcelブックです"
            f"（1シート＝1データセットの全{len(DATA_DICTIONARY)}シート＋目次「はじめに」／計{len(dict_df)}列の定義）。"
            "シート名は上の表の「列定義書のシート」と対応します。"
            "データを受け取った人が中身を解釈できるよう、CSVと一緒に配布してください。"
        )
        st.download_button(
            "列定義書をダウンロード（kumamoto_data_dictionary.xlsx）",
            to_dictionary_xlsx_bytes(dict_df, tuple(dict_notes)),
            file_name="kumamoto_data_dictionary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_dictionary",
        )
        # 定義書と実データのずれは異常ではなく一時的に起こりうるものなので、
        # 警告色ではなく理由の分かる備考として出す。
        for note in dict_notes:
            st.caption(note)
        with st.expander("列定義書をこの画面で見る"):
            # Excelのシート分けと同じ区切りで見られるようにタブにする
            dict_tabs = st.tabs([sheet for _, sheet, _, _ in DATA_DICTIONARY])
            # 「配布中CSVに存在」はxlsxと同じ条件（食い違いがあるときだけ）で出す
            drop_cols = ["csv_file", "dataset"]
            if not dictionary_has_missing_column(dict_df):
                drop_cols.append("in_actual_csv")
            for tab, (fname, _sheet, dataset, _cols) in zip(dict_tabs, DATA_DICTIONARY):
                with tab:
                    st.caption(f"{dataset}｜`{fname}`")
                    sub = dict_df[dict_df["csv_file"] == fname]
                    st.dataframe(
                        sub.drop(columns=drop_cols).rename(
                            columns={"in_actual_csv": "配布中CSVに存在"}
                        ),
                        use_container_width=True, hide_index=True,
                    )
        st.divider()

        st.caption(
            "出典を明記してご利用ください: 交通量は"
            "[JARTIC交通量オープンデータ](https://www.jartic-open-traffic.org/)、"
            "通行規制は「防災情報くまもと」の"
            "[通行規制情報](https://portal.bousai.pref.kumamoto.jp/?p=traffic)。"
            "通行規制の経路は始点・終点座標を[OSRM](https://project-osrm.org/)で"
            "道路網にスナップした推定値であり、元データそのものではありません。"
        )

    # ------------------------------------------------------------------
    # 道路規制×交通量タブ（通れる道マップの配布データ版・既定）
    # ------------------------------------------------------------------
    with tab_mlit:
        if post.empty:
            st.info("地震発生後のデータがまだありません。")
        else:
            render_mlit_beta_tab(
                point_summary, point_labels, quake_at, other_event_times,
                quake_info, mainshock, anomaly_end,
            )

    # ------------------------------------------------------------------
    # 人口推計値×交通量タブ（モバイル空間統計）
    # ------------------------------------------------------------------
    with tab_mesh:
        render_mesh_population_tab(
            point_summary, build_point_labels(point_summary), quake_at,
            mainshock, other_event_times,
        )

    # ------------------------------------------------------------------
    # 人口推計値の時点比較タブ（条件違いの2枚を並べる）
    # ------------------------------------------------------------------
    with tab_mesh_cmp:
        render_mesh_compare_tab(mainshock)


if __name__ == "__main__":
    main()
