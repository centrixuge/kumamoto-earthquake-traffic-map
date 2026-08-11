"""
「通れる道マップ」からダウンロードした各時点のZIPを、Excelの一覧にする。

どの時点にどのレイヤが入っているかは時点ごとにまちまちで
（初回は道路規制が無い、通行実績が5分割されている、など）、
ZIPを開かないと分からない。作業前に見渡せるよう表にしておく。

    python scripts/build_mlit_map_index.py

出力: data/mlit_r8kumamoto_map/通れる道マップ_データ一覧.xlsx
      シート「データ一覧」   … 時点ごとに何のレイヤが入っているか
      シート「規制情報の属性」… 道路規制情報の属性が、時点ごとに何件に入っているか
      シート「規制情報の型」  … 属性の組み合わせでレコードを分類したもの
"""
import glob
import json
import os
import re
import zipfile
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data", "mlit_r8kumamoto_map",
)
OUT_NAME = "通れる道マップ_データ一覧.xlsx"
FONT = "Meiryo"

# ZIP内のファイル名と、それが何の情報か。「通れる道マップ」の
# 表示レイヤに対応する（dourokisei=道路規制、tukoujisseki=通行実績）。
LAYERS = [
    ("道路規制情報", re.compile(r"^dourokisei\d*\.geojson$", re.I)),
    ("ETC2.0速度情報", re.compile(r"^ETC2\.0_speed_data\d*\.geojson$", re.I)),
    ("通行実績情報", re.compile(r"^tukoujisseki\d*\.geojson$", re.I)),
]

# 時刻がファイル名に入っていない回。配布元の notice.txt に
# 「260729data.zip is the data as of 2026/07/29 8:00」とある。
DATE_ONLY_HOUR = {"260729": "0800"}

# ファイルからは読み取れない、配布元の但し書き。
MANUAL_NOTES = {
    "2608071600data.zip": "規制情報のみ10:00時点",
}

# 道路規制情報のレコードは、属性の組み合わせが何通りかある。
# 名前は「どの属性を持つか」だけで付けている。データのどこにも出所は
# 書かれていないので、県が出したものか国が出したものかは分からない。
# どの道路種別が入るかは推測せず、実数を「道路種別の内訳」列に出す。
# 判定に使う目印の属性と、呼び名・説明。上から順に当てはめる。
RECORD_TYPES = [
    ("災害詳細つき", lambda k: "地整番号" in k,
     "住所・座標・開始日時に加えて、地整番号・整備局名・県番号・規制変更_日時/内容・"
     "迂回路/孤立集落/人身/物損/停電_有無を持つ"),
    ("方向・規制延長つき", lambda k: "規制方向" in k or "規制延長_Km" in k,
     "住所・座標・開始日時に加えて、規制方向と規制延長_Kmを持つ"),
    ("区間名のみ", lambda k: "始点" in k and "始点住所" not in k,
     "始点・終点が地点名（IC名など）。住所・座標・開始日時のいずれも持たない"),
    ("属性なし", lambda k: k <= {"name"},
     "整理IDだけで属性が無い線（紫・半透明）。何の規制かはデータから分からない"),
    ("住所・座標・日時", lambda k: True,
     "住所・緯度経度・規制開始_日時・延長_Kmを持つ、最も多い型"),
]

# 表示スタイルの属性（Leafletの描画用で、規制の中身ではない）
STYLE_KEYS_PREFIX = "_"


def parse_timestamp(stem: str) -> tuple:
    """ファイル名（YYMMDDhhmm）から日時を作る。戻り値は (datetime, 根拠)。"""
    if len(stem) == 10:
        return datetime.strptime(stem, "%y%m%d%H%M"), "ファイル名"
    if len(stem) == 6 and stem in DATE_ONLY_HOUR:
        return (
            datetime.strptime(stem + DATE_ONLY_HOUR[stem], "%y%m%d%H%M"),
            "notice.txt",
        )
    raise ValueError(f"日時を読み取れないファイル名: {stem}")


def read_regulations(zip_path: str) -> list:
    """ZIP内の道路規制情報のGeoJSONを読む（分割されていれば繋げる）。"""
    features = []
    with zipfile.ZipFile(zip_path) as z:
        for entry in z.namelist():
            if LAYERS[0][1].match(os.path.basename(entry)):
                # BOM付きで配布されているので utf-8-sig で読む
                data = json.loads(z.read(entry).decode("utf-8-sig"))
                features.extend(data.get("features", []))
    return features


def classify(keys: set) -> str:
    """属性の組み合わせから、レコードの型を決める。"""
    for label, matches, _ in RECORD_TYPES:
        if matches(keys):
            return label
    return RECORD_TYPES[-1][0]


def scan_regulations(rows: list, data_dir: str) -> None:
    """各時点の道路規制情報について、属性の出現件数と型の内訳を数える。"""
    for row in rows:
        if not row["layers"]["道路規制情報"]:
            row["reg"] = None
            continue
        features = read_regulations(os.path.join(data_dir, row["file"]))
        counts, types, geoms = Counter(), Counter(), Counter()
        roads = {}
        for feature in features:
            props = feature.get("properties", {})
            counts.update(props.keys())
            label = classify({
                k for k in props if not k.startswith(STYLE_KEYS_PREFIX)
            })
            types[label] += 1
            # 型ごとにどの道路種別が入っているかは、推測せず数える
            roads.setdefault(label, Counter())[
                props.get("道路種別") or "(値なし)"
            ] += 1
            geoms[(feature.get("geometry") or {}).get("type", "なし")] += 1
        row["reg"] = {
            "n": len(features), "keys": counts, "types": types,
            "roads": roads, "geoms": geoms,
        }


def scan(data_dir: str) -> list:
    rows = []
    for path in sorted(glob.glob(os.path.join(data_dir, "*.zip"))):
        name = os.path.basename(path)
        stem = name[: -len("data.zip")]
        stamp, basis = parse_timestamp(stem)
        with zipfile.ZipFile(path) as z:
            files = [
                os.path.basename(n) for n in z.namelist()
                if n.lower().endswith(".geojson")
            ]
        found, notes = {}, []
        for label, pattern in LAYERS:
            hits = [f for f in files if pattern.match(f)]
            found[label] = "〇" if hits else ""
            if len(hits) > 1:
                notes.append(f"{label}が{len(hits)}分割")
        if basis != "ファイル名":
            notes.append(f"時刻の出典: {basis}")
        if name in MANUAL_NOTES:
            notes.insert(0, MANUAL_NOTES[name])
        rows.append({
            "file": name,
            "stamp": stamp,
            "layers": found,
            "note": "、".join(notes),
        })
    rows.sort(key=lambda r: r["stamp"])
    scan_regulations(rows, data_dir)
    return rows


def build(rows: list, out_path: str) -> None:
    headers = ["ファイル名", "情報の日時および時刻"] + [n for n, _ in LAYERS] + ["備考"]
    wb = Workbook()
    ws = wb.active
    ws.title = "データ一覧"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="DDEBF7")

    for col, head in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=head)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for i, row in enumerate(rows, start=2):
        values = [row["file"], row["stamp"]] + [
            row["layers"][n] for n, _ in LAYERS
        ] + [row["note"]]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=value)
            c.font = Font(name=FONT, size=10)
            c.border = border
            if col == 2:
                c.number_format = "yyyy/mm/dd hh:mm"
                c.alignment = Alignment(horizontal="center")
            elif 3 <= col <= 2 + len(LAYERS):
                c.alignment = Alignment(horizontal="center")

    # 集計行。値ではなく数式で置き、行を足しても追随するようにする。
    total_row = len(rows) + 2
    label = ws.cell(row=total_row, column=1, value="〇の数")
    label.font = Font(name=FONT, bold=True, size=10)
    label.border = border
    ws.cell(row=total_row, column=2, value=f"全{len(rows)}ファイル").font = Font(
        name=FONT, bold=True, size=10
    )
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=2).border = border
    for col in range(3, 3 + len(LAYERS)):
        letter = get_column_letter(col)
        c = ws.cell(
            row=total_row, column=col,
            value=f'=COUNTIF({letter}2:{letter}{len(rows) + 1},"〇")',
        )
        c.font = Font(name=FONT, bold=True, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = border
    ws.cell(row=total_row, column=len(headers)).border = border

    # 出典と、日時をどう決めたかを表の下に残す（後から見て根拠が追えるように）
    notes = [
        "出典: 国土交通省「通れる道マップ」（https://www.mlit.go.jp/road/saigai/r8kumamoto/index.html）"
        "からダウンロードしたZIP。",
        "情報の日時および時刻: ファイル名の YYMMDDhhmm から読み取り（例 2608071600 → 2026/08/07 16:00）。"
        "260729data.zip だけ時刻が無く、配布元の notice.txt の記載（as of 2026/07/29 8:00）に従った。",
        "〇の判定: ZIP内のGeoJSONのファイル名で判定（dourokisei＝道路規制情報、"
        "ETC2.0_speed_data＝ETC2.0速度情報、tukoujisseki＝通行実績情報）。"
        "連番で分割されているものも1件として〇にし、分割数は備考に書いた。",
        "作成: scripts/build_mlit_map_index.py",
    ]
    for j, text in enumerate(notes):
        c = ws.cell(row=total_row + 2 + j, column=1, value=text)
        c.font = Font(name=FONT, size=9, color="595959")
        c.alignment = Alignment(vertical="top")

    widths = [22, 20] + [15] * len(LAYERS) + [30]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    build_attribute_sheet(wb, rows, border, head_fill)
    build_type_sheet(wb, rows, border, head_fill)

    wb.save(out_path)


def _reg_rows(rows: list) -> list:
    return [r for r in rows if r.get("reg")]


def _matrix_header(ws, rows: list, first_label: str, border, head_fill,
                   extra: list = ()) -> None:
    """1列目が項目名、2列目以降が時点、という共通の見出しを作る。"""
    heads = [first_label] + list(extra) + [
        f'{r["stamp"]:%m/%d %H:%M}' for r in rows
    ]
    for col, head in enumerate(heads, start=1):
        c = ws.cell(row=1, column=col, value=head)
        c.font = Font(name=FONT, bold=True, size=9)
        c.fill = head_fill
        c.alignment = Alignment(
            horizontal="center", vertical="bottom",
            textRotation=0 if col <= 1 + len(extra) else 90,
        )
        c.border = border


def build_attribute_sheet(wb, rows: list, border, head_fill) -> None:
    """属性 × 時点。値はその属性を持つ地物の件数。"""
    rows = _reg_rows(rows)
    ws = wb.create_sheet("規制情報の属性")
    # 属性の並びは、出てくる時点が多い順（＝共通して使えるものが上）
    order = Counter()
    for r in rows:
        order.update(r["reg"]["keys"].keys())
    keys = sorted(
        order, key=lambda k: (k.startswith(STYLE_KEYS_PREFIX), -order[k], k)
    )
    _matrix_header(ws, rows, "属性", border, head_fill, ["区分", "出現時点数"])

    def _put(r_i, c_i, value, bold=False, color=None):
        c = ws.cell(row=r_i, column=c_i, value=value)
        c.font = Font(name=FONT, size=9, bold=bold, color=color)
        c.border = border
        if c_i > 1:
            c.alignment = Alignment(horizontal="center")
        return c

    line = 2
    _put(line, 1, "地物数（合計）", bold=True)
    _put(line, 2, "", bold=True)
    _put(line, 3, len(rows), bold=True)
    for j, r in enumerate(rows):
        _put(line, 4 + j, r["reg"]["n"], bold=True)
    line += 1
    for key in keys:
        style = key.startswith(STYLE_KEYS_PREFIX)
        _put(line, 1, key, color="808080" if style else None)
        _put(line, 2, "表示スタイル" if style else "属性",
             color="808080" if style else None)
        _put(line, 3, order[key])
        for j, r in enumerate(rows):
            n_all, n_key = r["reg"]["n"], r["reg"]["keys"].get(key, 0)
            if not n_key:
                _put(line, 4 + j, "－", color="BFBFBF")
            elif n_key == n_all:
                _put(line, 4 + j, "全")
            else:
                _put(line, 4 + j, n_key, color="C00000")
        line += 1

    notes = [
        "セルの意味: 「全」＝その時点の全地物が持つ／数字＝持っている地物の数"
        "（赤字。持たない地物が混ざる）／「－」＝その時点には無い。",
        "アンダースコアで始まる属性は地図の描画用（色・太さなど）で、規制の中身ではない。",
        "属性が地物ごとに違うのは、レコードの型が混ざっているため。内訳は「規制情報の型」シート。",
    ]
    for j, text in enumerate(notes):
        c = ws.cell(row=line + 1 + j, column=1, value=text)
        c.font = Font(name=FONT, size=9, color="595959")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    for j in range(len(rows)):
        ws.column_dimensions[get_column_letter(4 + j)].width = 6
    ws.freeze_panes = "D2"


def build_type_sheet(wb, rows: list, border, head_fill) -> None:
    """レコードの型 × 時点。値はその型の地物の件数。"""
    rows = _reg_rows(rows)
    ws = wb.create_sheet("規制情報の型")
    _matrix_header(
        ws, rows, "レコードの型", border, head_fill,
        ["持っている属性", "道路種別の内訳（全時点の合計）"],
    )

    # 型ごとの道路種別は、全時点を合算した実数を出す
    totals = {}
    for r in rows:
        for label, counter in r["reg"]["roads"].items():
            totals.setdefault(label, Counter()).update(counter)

    line = 2
    for label, _, desc in RECORD_TYPES:
        c = ws.cell(row=line, column=1, value=label)
        c.font = Font(name=FONT, size=9, bold=True)
        c.border = border
        c.alignment = Alignment(vertical="center")
        c = ws.cell(row=line, column=2, value=desc)
        c.font = Font(name=FONT, size=9)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="center")
        breakdown = "、".join(
            f"{k} {v}" for k, v in totals.get(label, Counter()).most_common()
        )
        c = ws.cell(row=line, column=3, value=breakdown or "－")
        c.font = Font(name=FONT, size=9)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="center")
        for j, r in enumerate(rows):
            n = r["reg"]["types"].get(label, 0)
            c = ws.cell(row=line, column=4 + j, value=n if n else "－")
            c.font = Font(name=FONT, size=9, color=None if n else "BFBFBF")
            c.alignment = Alignment(horizontal="center")
            c.border = border
        line += 1

    # 合計は数式にして、型を足したときに自動で合うようにする
    c = ws.cell(row=line, column=1, value="合計")
    c.font = Font(name=FONT, size=9, bold=True)
    c.border = border
    for col in (2, 3):
        ws.cell(row=line, column=col).border = border
    for j in range(len(rows)):
        letter = get_column_letter(4 + j)
        c = ws.cell(
            row=line, column=4 + j,
            value=f"=SUM({letter}2:{letter}{line - 1})",
        )
        c.font = Font(name=FONT, size=9, bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = border
    line += 1

    c = ws.cell(row=line, column=1, value="ジオメトリ")
    c.font = Font(name=FONT, size=9, bold=True)
    c.border = border
    for col in (2, 3):
        ws.cell(row=line, column=col).border = border
    for j, r in enumerate(rows):
        geoms = r["reg"]["geoms"]
        text = "線" if set(geoms) == {"LineString"} else "＋".join(
            f'{"線" if g == "LineString" else "点" if g == "Point" else g}{n}'
            for g, n in geoms.items()
        )
        c = ws.cell(row=line, column=4 + j, value=text)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(horizontal="center")
        c.border = border

    notes = [
        "型の名前は「どの属性を持つか」だけで付けている。データのどこにも出所は"
        "書かれていないので、県が出したものか国が出したものかは分からない。"
        "どの道路種別が入るかも推測せず、実数をC列に出している。",
        "判定は上から順に当てはめる（地整番号があれば「災害詳細つき」、"
        "無くて規制方向があれば「方向・規制延長つき」…）。",
        "「区間名のみ」は座標も規制開始日時も持たないため、線形と期間はこのデータからは作れない"
        "（ジオメトリの線はあるので、地図に出すことはできる）。",
        "「属性なし」は整理IDだけの紫の線で、07/31〜08/01の3時点にだけ現れる。",
        "「災害詳細つき」は08/04以降の3件だけで、いずれも道路種別は補助国道。",
        "道路種別の値そのものにも揺れがある（「一般国道」と「補助国道」「直轄国道」が混在、"
        "路線名が入っている1件など）。",
    ]
    for j, text in enumerate(notes):
        c = ws.cell(row=line + 2 + j, column=1, value=text)
        c.font = Font(name=FONT, size=9, color="595959")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 40
    for j in range(len(rows)):
        ws.column_dimensions[get_column_letter(4 + j)].width = 6
    ws.freeze_panes = "D2"


def main() -> None:
    rows = scan(DATA_DIR)
    out_path = os.path.join(DATA_DIR, OUT_NAME)
    build(rows, out_path)
    for row in rows:
        flags = "".join(
            (row["layers"][n] or "-").ljust(2) for n, _ in LAYERS
        )
        print(f'{row["file"]:<22} {row["stamp"]:%Y-%m-%d %H:%M}  {flags} {row["note"]}')
    print(f"\n{len(rows)}件 -> {out_path}")


if __name__ == "__main__":
    main()
